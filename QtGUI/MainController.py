import sys
from PyQt5 import QtGui
from PyQt5 import QtCore
#from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, 
#                             QToolTip, QMessageBox, QLabel,QDialog,QTableWidget,QTableWidgetItem,
#                             QHeaderView)

from PyQt5.QtWidgets import *

import datetime
import decimal
import math
from fractions import Fraction
import threading
import time

#Sử dụng openVPN để xử lý trường hợp truy cập public
#import fetchip
#import socket
#import openvpn_api.vpn
#import openvpn_status
#import netaddr



from MainWin import *


from M1_qlhsPop import *
from M11_locdiemdanhPop import *
from M12_tbwthemhocsinh import *
from M13_xoahocsinh import *
from M14_thongkediemdanh import *
from M16_danhsachhocsinh import *

from M2_qlgv import *
from M21_dsgv import *
from M22_themgv import *
from M23_phanconggv import *
from M24_xoagv import *

from M3_quanlyhocphi import *
from M31_xoahocphidmcu import *
from M32_themhpdinhmuc import *
from M33_ghinhantienhoc import *

from M4_phanmonnhomtkb import *
from M41_lkthemxoanhom import *
from M42_thaydoihstrongnhom import *
from M43_thaydoigiohoc import *

from M5_phantichloinhuan import *
from M51_bangchiluonggv import *
from M52_capnhatthongtinluong import *
from M53_luuchitraluonggv import *

from M6_quanlybaihocgiohoc import *

from M7_thongtinlienheonline import *
from M71_txcninfophuhuynh import *
from M72_guitncanhan import *
from M73_guitnnhom import *

#from splashScreen import *











#Import for Database Connection
import paramiko
import os
import pymysql
import pandas as pd
from paramiko import SSHClient
from sshtunnel import SSHTunnelForwarder
from os.path import expanduser

#Import for fbchat Connection

#from fbchat import Client
#from fbchat.models import *

#Get fbID
#import requests

#Import for excel & csv controller
from openpyxl import Workbook
from openpyxl.utils import get_column_letter







#This part intended left blank

#OpenVPN Handler Modules truy cập thông qua DDNS của NoIP sử dụng noipy package
#def ddns_noip:
    



##Data loading module
#class loading_windows(LoadingGif):
    
#    #Khởi tạo
#    def setupUi(self,windows_loading):
#        super().mainUi(windows_loading)
#        self.ui = ()
#        self.ui.mainUi(windows_loading)
#        #Hiển thị view
#        windows_loading.show()

#    #Đóng
#    def closeUi(self,windows_loading):
#        super().setupUi(windows_loading)
#        #Tắt View
#        windows_loading.hide()




# Database Handler Modules
def run_cmd(command):
    

    ##Mở cửa sổ loading
    #self.ui = loading_windows()
    #self.ui.mainUi(windows_loading)
    
    #Connect to Remote Server
    k = paramiko.RSAKey.from_private_key_file("D:\CodingProjects\MyManageProject\kvmanagement\id_rsa_server",password = "kvdatabaselockdown")
    
    #Đổi option key cho file cài đặt
    #k = paramiko.RSAKey.from_private_key_file("C:/Program Files (x86)/khanhvuedu/id_rsa_server",password = "kvdatabaselockdown")
    
    #Old codes used - Just for rethinking
    '''

    remoteComp = paramiko.SSHClient()
    remoteComp.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    print("Connecting...")
    remoteComp.connect(hostname = "192.168.2.117", username = "longpham", pkey = k)

    '''


    sql_hostname = '127.0.0.1'
    sql_username = 'root'
    sql_password = '0377122966longpham!'
    sql_main_database = 'kvdatabase'
    sql_port = 3306
    ssh_host = '192.168.2.117'
    ssh_user = 'longpham'
    ssh_port = 22
    sql_ip = '1.1.1.1.1'

    with SSHTunnelForwarder(
            (ssh_host,ssh_port),ssh_username = ssh_user,
            ssh_pkey=k,
            remote_bind_address = (sql_hostname,sql_port)) as tunnel:
        conn = pymysql.connect(host='127.0.0.1',user=sql_username,
                         passwd=sql_password,db=sql_main_database,
                         port=tunnel.local_bind_port)
        
        #Counter start
        #startcounter = datetime.datetime.now()
        #microsec_start = startcounter.microsecond

        # Check if database is connected Command - obmit if Connection establish fine at first check!!
        print('Kvdatabase connected!!!')
        cursor = conn.cursor()



        #Test commands establish to database!! Obmit if command establish fine!!
        cursor.execute(command)
        if (command.find("SELECT")==-1): 
            conn.commit()
        if (command.find("SELECT") != -1):
            records = cursor.fetchall()
        
            #Output dữ liệu vào biến toàn cục recs, type list
            global recs
            recs = []
            recs = list(records)
        
        cursor.close()
        #Close ssh connection
        conn.close()

        #windows_loading.hide()

        #Counter end and output
        #endcounter = datetime.datetime.now()
        #microsec_end = endcounter.microsecond
        #process_time = endcounter - startcounter
        #microsec_process = microsec_end - microsec_start
        #print('Connected in ',process_time ,' secs and ',microsec_process , ' microseconds!')













#This part intended left blank


# Window 0 - Cửa sổ làm việc chính
class MainWinOptimized(Ui_MainWindow):


    def setupUi(self,MainWinOpt):
        super().setupUi(MainWinOpt)
        self.qlhs.clicked.connect(self.Open_qlhsPop)
        self.qlgv.clicked.connect(self.Open_themgv)
        self.qlhp.clicked.connect(self.Open_qlhp)
        self.pmvn.clicked.connect(self.Open_pmntkb)
        self.ptln.clicked.connect(self.Open_ptln)
        
        self.qlbhgh.setEnabled(False)
        self.qlbhgh.clicked.connect(self.Open_qlbhgh)
        
        self.stuconnect.setEnabled(False)
        self.stuconnect.clicked.connect(self.Open_ttlhph)


    #Open Window 1 - Cửa sổ quản lý học sinh
    def Open_qlhsPop(self):
        self.ui = OpWinM1_qlhs()
        self.ui.setupUi(window_1)
        MainWinOpt.hide()
   
    #Open Window 2 - Cửa sổ quản lý giáo viên
    def Open_themgv(self):
        self.ui = OpWinM2_qlgv()
        self.ui.setupUi(window_2)
        MainWinOpt.hide()

    #Open Window 3 - Cửa sổ quản lý học phí
    def Open_qlhp(self):
        self.ui = OpWinM3_qlhp()
        self.ui.setupUi(window_3)
        MainWinOpt.hide()

    #Open Window 4 - Cửa sổ phân môn nhóm và thời khóa biểu
    def Open_pmntkb(self):
        self.ui = OpWinM4_pmntkb()
        self.ui.setupUi(window_4)
        MainWinOpt.hide()


    #Open Window 5 - Cửa sổ phân tích lợi nhuận
    def Open_ptln(self):
        self.ui = OpWinM5_ptln()
        self.ui.setupUi(window_5)
        MainWinOpt.hide()

    #Open Window 6 - Cửa sổ quản lý bài học và giờ học
    def Open_qlbhgh(self):
        self.ui = OpWinM6_qlbhgh()
        self.ui.setupUi(window_6)
        MainWinOpt.hide()

    #Open Window 7 - Cửa sổ quản lý thông tin và liên hệ phụ huynh online
    def Open_ttlhph(self):
        self.ui = OpWinM7_ttphonline()
        self.ui.setupUi(window_7)
        MainWinOpt.hide()










#This part intended left blank


# Window 1 - Dialog quản lý học sinh
class OpWinM1_qlhs(Ui_qlhsPop):

    # Open Window 1-1 - Dialog lọc để điểm danh
    def Open_locdiemdanhPop(self):
        self.ui = OpWinM11_confirmldd()
        self.ui.setupUi(window_11)
        window_1.hide()

    # Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_1.hide()

    # Open Window 1-2 - Dialog chọn số học sinh thêm mới
    def Open_sohsthemmoi(self):
        self.ui = OpWinM12_confirmnewstu()
        self.ui.setupUi(window_12)
        window_1.hide()

    # Open Window 1-3 - Dialog xóa học sinh đã nghỉ
    def Open_xoahs(self):
        self.ui = OpWinM13_xoahocsinh()
        self.ui.setupUi(window_13)
        window_1.hide()


    # Open Window 1-4 - Dialog thống kê điểm danh học sinh
    def Open_tkddhs(self):
        self.ui = OpWinM14_thongkediemdanh()
        self.ui.setupUi(window_14)
        window_1.hide()

    #Open Window 1-6 - Dialog danh sách học sinh
    def Open_dshs(self):
        self.ui = OpWinM16_danhsachhocsinh()
        self.ui.setupUi(window_16)
        window_1.hide()



    #Khởi tạo
    def setupUi(self,window_1):
        super().setupUi(window_1)
        
        #Setup starting status
        self.tbvn.setEnabled(False)

        #Buttons clicks definitions 
        self.ldd.clicked.connect(self.Open_locdiemdanhPop)
        self.thsm.clicked.connect(self.Open_sohsthemmoi)
        self.xoahs.clicked.connect(self.Open_xoahs)
        self.bdd.clicked.connect(self.Open_tkddhs)
        self.dshs.clicked.connect(self.Open_dshs)

        self.quaylai.clicked.connect(self.ql)

        #Show current defined window
        window_1.show()



#Window 1-1 - Dialog lọc để điểm danh
class OpWinM11_confirmldd(Ui_locdiemdanh):
    def __init__(self):
        #Def Translate variable
        _translate = QtCore.QCoreApplication.translate
        
        #Khởi tạo danh sách giáo viên
        cmd_dsgv = 'SELECT * FROM dsgiaovien;'
        run_cmd(cmd_dsgv)
        self.dsgv = ['0',]
        for row in recs:
            self.dsgv.append(row[0]+' - '+row[1])
        #for row in self.dsgv:
        #    print(row)
        #i=0
        #for row in dsgv:
        #    i +=1 
        #    self.combo_giaovien.setItemText(i,_translate("locdiemdanh",row))

        #Chèn vào combobox giáo viên - Đang gặp  lỗi vì khai báo trước khi combo_giaovien được khai báo!!!
        #i=0
        #for row in dsgv:
        #    i +=1 
        #    self.combo_giaovien.setItemText(i,_translate("locdiemdanh",row))
        
        self.dsnhom = ['0',]
        cmd_dsnhom = 'SELECT * FROM dsnhom;'
        run_cmd(cmd_dsnhom)
        for row in recs:
            self.dsnhom.append(row[0]+' - '+row[1])
        #for row in self.dsnhom:
        #    print(row)

        


    #Khởi tạo chính
    def setupUi(self,window_11):
        super().setupUi(window_11)


        #Show current defined window
        window_11.show()

        #Up dữ liệu lên combobox
        self.combo_giaovien.addItems(self.dsgv)
        self.combo_nhom.addItems(self.dsnhom)

        #Khởi tạo giá trị mặc định cho ngày điểm danh
        self.ngaydd.setDate(QtCore.QDate.currentDate())

        #Def nút hoàn tất
        self.hoantat.clicked.connect(self.but_hoantat)
        self.loc.clicked.connect(self.but_loc)
        self.xndiemdanh.clicked.connect(self.but_xndiemdanh)
        self.xoadiemdanh.clicked.connect(self.but_xnxoadd)
        

    #Def nút hoàn tất
    def but_hoantat(self):
        window_1.show()
        window_11.hide()


    #Def but loc
    def but_loc(self):
        #Reset bảng lại trước khi xử lý
        self.diemdanh_view.clear
        
        #Lấy kết quả tùy chọn 
        DelCheckbox = QTableWidgetItem()
        DelCheckbox.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                
        DelCheckbox = self.cb_xoadiemdanh.checkState()
        kt = int(DelCheckbox)

        
        if kt:
            #Lệnh nền để khởi tạo tiêu đề bảng hiển thị
            _translate = QtCore.QCoreApplication.translate
            #Chèn thông báo
            #self.label_4.setText(_translate("locdiemdanh",""))
            #self.label_4.setText(_translate("locdiemdanh", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Tùy chọn để xóa</span></p></body></html>"))
            # Xử lý xóa các điểm danh bị nhầm lẫn
            op_ngaydd = self.ngaydd.date().toString('yyyy-MM-dd')
            wdop_ngaydd = self.ngaydd.date().dayOfWeek() - 1
            #print(wdop_ngaydd)

            op_ngaydd = self.ngaydd.date().toPyDate()
            thudd = op_ngaydd.weekday()
            print(thudd)

            cmd_sort = 'SELECT diemdanhhs.IDDiemDanh,dshocsinh.HoTenHS,diemdanhhs.MaHS,diemdanhhs.MaMonHoc,diemdanhhs.MaNhom,diemdanhhs.NgayGioHoc,thoikhoabieu.NgayGioHoc,thoikhoabieu.Thu,dsgiaovien.tenGV FROM diemdanhhs INNER JOIN dshocsinh ON dshocsinh.MaHS = diemdanhhs.MaHS INNER JOIN thoikhoabieu ON diemdanhhs.MaNhom = thoikhoabieu.MaNhom INNER JOIN dsnhom ON dsnhom.MaNhom = diemdanhhs.MaNhom INNER JOIN dsgiaovien ON dsgiaovien.maGV = dsnhom.maGV WHERE (thoikhoabieu.TinhTrangLich = 1) AND (DATE(diemdanhhs.NgayGioHoc) = "' + str(op_ngaydd) + '") AND (thoikhoabieu.Thu = ' + str(thudd) + ');'
            print(cmd_sort)
            run_cmd(cmd_sort)
            
            self.datafordel = []
            #print(recs)
            for row in recs:
                per_row = []
                for item in row:
                    per_row.append(item)
                self.datafordel.append(per_row)

            print(self.datafordel)

            #for row in recs:
            #    print(row)

            #Xây dựng bảng để output dữ liệu
            self.diemdanh_view.setRowCount(len(recs))
            self.diemdanh_view.setColumnCount(10)
            self.header = self.diemdanh_view.horizontalHeader()      
            for col in range(0,10):
                self.header.setSectionResizeMode(col, QHeaderView.ResizeToContents)

            
            #Upload dữ liệu lên bảng hiển thị diemdanh_view
            line = -1
            for row in recs:
                line += 1
                self.diemdanh_view.setItem(line,0,QTableWidgetItem(row[2]))
                self.diemdanh_view.setItem(line,1,QTableWidgetItem(row[1]))
                self.diemdanh_view.setItem(line,2,QTableWidgetItem(row[4]))
                
                dt = datetime.datetime.strptime(str(row[5]),'%Y-%m-%d %H:%M:%S')
                #print(dt.date())
                #print(dt.time())

                op_thu = self.chg_wd(dt.weekday())
                tkb_thu = self.combo_buoihoc.currentText()
                #print(tkb_thu)
                #print(op_thu)
                #if op_thu == tkb_thu:
                #    print("OK!!!")
                #else:
                #    print('Not compatible!!')


                #Thêm xử lý tình trạng khi thứ điểm danh được chọn












                self.diemdanh_view.setItem(line,3,QTableWidgetItem(op_thu))

                
                self.diemdanh_view.setItem(line,4,QTableWidgetItem(str(op_ngaydd)))
                if tkb_thu == '0':
                    self.diemdanh_view.setItem(line,5,QTableWidgetItem("Không dùng!"))
                else:
                    self.diemdanh_view.setItem(line,5,QTableWidgetItem(str(tkb_thu)))

                op_time = str(dt.time())
                #print(op_time,'-',str(type(op_time)))
                self.diemdanh_view.setItem(line,6,QTableWidgetItem(op_time))
                self.diemdanh_view.setItem(line,7,QTableWidgetItem(row[8]))
                self.diemdanh_view.setItem(line,8,QTableWidgetItem("Đã học!!"))

                #Xử lý checkbox cột kế cuối
                chkBoxItem = QTableWidgetItem()
                chkBoxItem.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                chkBoxItem.setCheckState(QtCore.Qt.Unchecked)       
                self.diemdanh_view.setItem(line,9,QTableWidgetItem(str(row[0])))
                

                #Xử lý tác vụ xóa đã được thực hiện bởi button Xóa điểm danh nhầm



        else:
            #Lệnh nền để khởi tạo tiêu đề bảng hiển thị
            _translate = QtCore.QCoreApplication.translate
            #Tạo lệnh lọc
            sort_ip = [str(self.combo_lop.currentText()),str(self.combo_giaovien.currentText()),str(self.combo_nhom.currentText())
                   ,str(self.combo_buoihoc.currentText()),self.giohoc.value(),self.phuthoc.value()]
            #print(sort_ip)

            cmd_sort = 'SELECT thoikhoabieu.IDTKB,hstheonhom.MaNhom,hstheonhom.MaHS,dshocsinh.HoTenHS,dshocsinh.Lop,thoikhoabieu.NgayGioHoc,dshocsinh.TinhTrangHS,dsgiaovien.tenGV,thoikhoabieu.Thu FROM thoikhoabieu INNER JOIN hstheonhom ON thoikhoabieu.MaNhom = hstheonhom.MaNhom INNER JOIN dsnhom ON thoikhoabieu.MaNhom = dsnhom.MaNhom JOIN dshocsinh ON hstheonhom.MaHS = dshocsinh.MaHS INNER JOIN dsgiaovien ON dsnhom.maGV = dsgiaovien.maGV WHERE (dshocsinh.TinhTrangHS = 1) AND '
        
            add_cmdstr = ';'
            #print(cmd_sort)

            if sort_ip[0] != '0':
                add_cmdstr = 'AND dshocsinh.Lop = ' + sort_ip[0] + add_cmdstr

            if sort_ip[1] != '0':
                add_cmdstr = 'AND dsgiaovien.maGV = "' + sort_ip[1][:4] + '" ' + add_cmdstr
        
            if sort_ip[2] != '0':
                add_cmdstr = 'AND hstheonhom.MaNhom = "' + sort_ip[2][:12] + '" ' + add_cmdstr
        
            if sort_ip[3] != '0':
                kq_thu = self.chg_date(sort_ip[3])
                add_cmdstr = 'AND thoikhoabieu.Thu = ' + str(kq_thu) + ' ' + add_cmdstr

            if ((sort_ip[4] != 0) or (sort_ip[5]!=0)):
                gio = 'time("' +str(sort_ip[4]) + ':' + str(sort_ip[5]) + ':' + '00") <= '
                dieukien = gio + 'time(thoikhoabieu.NgayGioHoc)'
                add_cmdstr = 'AND ' + dieukien + ' ' + add_cmdstr

           

            
            if add_cmdstr[0] != ';':
                add_cmdstr = add_cmdstr[4:]
            #print(add_cmdstr)
            if add_cmdstr[0] ==';':
                cmd_sort = cmd_sort[:-4]
            cmd_sort += add_cmdstr
        
            #Chèn thông báo
            #self.label_4.setText(_translate("locdiemdanh",""))
            #self.label_4.setText(_translate("locdiemdanh", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Lọc đã hoàn tất!!!</span></p></body></html>"))
            
            print(cmd_sort)

            #Push lệnh lọc đã tạo lên database truy xuất
            run_cmd(cmd_sort)
            #for row in recs:
            #   print(row)

            #Xây dựng bảng để output dữ liệu
            self.diemdanh_view.setRowCount(len(recs))
            self.diemdanh_view.setColumnCount(10)
            self.header = self.diemdanh_view.horizontalHeader()  
            
            for col in range(0,10):
                self.header.setSectionResizeMode(col, QHeaderView.ResizeToContents)


            line = -1
            for row in recs:
                line += 1
                self.diemdanh_view.setItem(line,0,QTableWidgetItem(row[2]))
                self.diemdanh_view.setItem(line,1,QTableWidgetItem(row[3]))
                self.diemdanh_view.setItem(line,2,QTableWidgetItem(row[1]))
            
                #Xử lý datetime
                dt = datetime.datetime.strptime(str(row[5]),'%Y-%m-%d %H:%M:%S')
                #print(dt.date())
                #print(dt.time())

                op_thu = self.chg_wd(dt.weekday())
                #print(op_thu)
                self.diemdanh_view.setItem(line,3,QTableWidgetItem(op_thu))
            
                op_time = str(dt.time())
                #print(op_time,'-',str(type(op_time)))
                self.diemdanh_view.setItem(line,6,QTableWidgetItem(op_time))

                self.diemdanh_view.setItem(line,7,QTableWidgetItem(row[7]))

                #Xử lý chèn ngày từ option ngày
                op_ngaydd = self.ngaydd.date().toPyDate()
                #print(op_ngaydd,' - ',type(op_ngaydd))

                self.diemdanh_view.setItem(line,4,QTableWidgetItem(str(op_ngaydd)))
                #Ghi nhận thứ điểm danh
                self.diemdanh_view.setItem(line,5,QTableWidgetItem(self.chg_wd(op_ngaydd.weekday())))

                #Xử lý checkbox cột kế cuối
                chkBoxItem = QTableWidgetItem()
                chkBoxItem.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                chkBoxItem.setCheckState(QtCore.Qt.Unchecked)       
                self.diemdanh_view.setItem(line,8,chkBoxItem)

                self.diemdanh_view.setItem(line,9,QTableWidgetItem("Không dùng!!"))


    # Def đổi thứ cho but_loc
    def chg_date(self,ipstring):
        if ipstring == 'Thứ hai':
            return 0
        if ipstring == 'Thứ ba':
            return 1
        if ipstring == 'Thứ tư':
            return 2
        if ipstring == 'Thứ năm':
            return 3
        if ipstring == 'Thứ sáu':
            return 4
        if ipstring == 'Thứ bảy':
            return 5
        if ipstring == 'Chủ nhật':
            return 6

    #Def đổi weekday thành thứ
    def chg_wd(self,ipint):
        if ipint == 0:
            return 'Thứ hai'
        if ipint == 1:
            return 'Thứ ba'
        if ipint == 2:
            return 'Thứ tư'
        if ipint == 3:
            return 'Thứ năm'
        if ipint == 4:
            return 'Thứ sáu'
        if ipint == 5:
            return 'Thứ bảy'
        if ipint == 6:
            return 'Chủ nhật'

    # Def but xác nhận điểm danh
    def but_xndiemdanh(self):
        
        #Lệnh nền để khởi tạo tiêu đề bảng hiển thị
        _translate = QtCore.QCoreApplication.translate

        #Extract data
        nrow = self.diemdanh_view.rowCount()
        ncol = self.diemdanh_view.columnCount()
        per_rec = []
        alldata = []

        #print(str(nrow),' - ',str(ncol))
        

        for row in range(0,nrow):
            for col in range(0,8):
                item = str(self.diemdanh_view.item(row,col).text())
                per_rec.append(item)

            chkBoxItem = QTableWidgetItem()
            chkBoxItem.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                
            chkBoxItem = self.diemdanh_view.item(row,8).checkState()
            #print(str(chkBoxItem),'-',type(chkBoxItem))
            per_rec.append(int(chkBoxItem))

            for col in range(9,10):
                item = str(self.diemdanh_view.item(row,col).text())
                per_rec.append(item)

            alldata.append(per_rec)
            per_rec = []

        for row in alldata:
            print(row)
        
        #Chèn dữ liệu vào database
        #Lấy max IDDiemDanh
        cmd_ipId = 'SELECT MAX(IDDiemDanh) FROM diemdanhhs;'
        run_cmd(cmd_ipId)

        #Khởi tạo chèn dữ liệu
        maxId = 0
        for row in recs:
            print(row)
        if type(recs[0][0]) == int:
            maxId = recs[0][0]
        print(maxId)

        idIp = maxId

        for row in alldata:
            #print(row)
            if row[8]:
                idIp += 1
                #Cập nhật record điểm danh
                #print(row)
                #print(idIp)
                dtIp = row[4] + ' ' + row[6]
                cmd_diemdanh = 'INSERT INTO diemdanhhs (IDDiemDanh,MaHS,MaMonHoc,MaNhom,NgayGioHoc,dmmonhoc_MaMonHoc,dshocsinh_MaHS) VALUES ('
                cmd_diemdanh = cmd_diemdanh + str(idIp) + ', "' + row[0] + '", "' + row[2][:4] + '", "' + row[2] + '", "' + dtIp + '", "' + row[2][:4] + '", "' + row[0] +'");'

                run_cmd(cmd_diemdanh)
                print(cmd_diemdanh)
        
                #Đẩy counter trong defhocphi để ghi nhận tính học phí
                cmd_laymagv = 'SELECT maGV FROM dsnhom WHERE MaNhom = "' + row[2] + '";'
                #print(cmd_laymagv)
                run_cmd(cmd_laymagv)
                #print(recs[0][0])
                cmd_updatedefhp = 'UPDATE defhocphi SET SoBuoiDaHoc = SoBuoiDaHoc + 1 WHERE MaHS = "' + row[0] + '" AND MaMonHoc = "' + row[2][:4] + '" AND MaGV = "' + recs[0][0] + '";'
                #print(cmd_updatedefhp)
                run_cmd(cmd_updatedefhp)



        #Xác nhận điểm danh xong
        #self.label_4.setText(_translate("locdiemdanh", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt;\">Đã điểm danh xong!!!</span></p></body></html>"))


    #Def nút xác nhận xóa điểm danh
    def but_xnxoadd(self):
        ipcodestr = self.ipidxoa.text()
        ipcd = ipcodestr.split(',')
        for row in ipcd:
            
            #Chọn mã học sinh (2) , Mã Môn học (3) , Mã nhóm (4) để lọc id cập nhật thông tin vào defhocphi từ biến self.datafordel
            counter = 0
            while str(self.datafordel[counter][0]) != row:
                counter += 1
            
            print(self.datafordel[counter])
            #Lọc mã giáo viên để update
            cmd_slmagv = 'SELECT maGV FROM dsnhom WHERE MaNhom = "' + self.datafordel[counter][4] + '";'
            #print(cmd_slmagv)
            run_cmd(cmd_slmagv)
            print(recs[0][0])
            magv_del = recs[0][0]

            cmd_updatedefhp = 'UPDATE defhocphi SET SoBuoiDaHoc = SoBuoiDaHoc - 1 WHERE MaHS = "' + self.datafordel[counter][2] + '" AND MaMonHoc = "' + self.datafordel[counter][3] + '" AND MaGV = "' + magv_del + '";'
            print(cmd_updatedefhp)
            run_cmd(cmd_updatedefhp)
            cmd_del = 'DELETE FROM diemdanhhs WHERE IDDiemDanh = ' + row + ';'
            run_cmd(cmd_del)


#Window 1-2 - Dialog thực thi việc thêm học sinh mới
class OpWinM12_confirmnewstu(Ui_tbwthemhocsinhmoi):
    #Khởi tạo
    def setupUi(self,window_12):
        super().setupUi(window_12)

        #button Hoàn tất
        self.hoantat.clicked.connect(self.but_hoantat)

        #button Xác nhận
        self.xacnhan.clicked.connect(self.but_xacnhan)
        
        #button Cập nhật học sinh mới vào hệ thống
        self.updatestu.clicked.connect(self.but_updatestu)

        #Show current defined window
        window_12.show()

        
    #Def Nút hoàn tất
    def but_hoantat(self):
        window_1.show()
        window_12.hide()

    #Mở Dialog thêm học sinh mới
    def but_xacnhan(self):
        w12_sohs = self.sohs.value()
        self.tbwthemhs.setRowCount(w12_sohs)
        self.tbwthemhs.setColumnCount(6)
        self.header = self.tbwthemhs.horizontalHeader()       
        self.header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  
        self.header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        
        
        #Chọn Mã học sinh để tạo mã mới
        sortMaHS = "SELECT MaHS FROM dshocsinh;"
        run_cmd(sortMaHS)
        ttma = []
        for row in recs:
            ttma.append(int(row[0][1:]))
        cs = max(ttma)

        #Xuất mã mới vào bảng
        for self.row in range(0,w12_sohs):
            cs += 1
            code_created = "S0" + str(cs)
            self.tbwthemhs.setItem(self.row,5,QTableWidgetItem(code_created))
        print("CMD Finished!!!")
        #Test append database
        #cmd1 = 'SELECT * FROM kvdatabase.dshocsinh;'
        #run_cmd(cmd1)
        #for row in recs:
        #    print(row)
        #print('CMD1 Finished!!')

        #cmd2 = 'INSERT INTO kvdatabase.dshocsinh(MaHS,HoTenHS,NamSinh,NgayVaoHoc,Lop,SoDienThoai,TinhTrangHS) VALUES("S123","Zeta",1995,"2020-03-05",9,"0377122966",1);'
        #run_cmd(cmd2)
        #print('CMD2 -ADDED Finished!!')


        #cmd3 = 'SELECT * FROM kvdatabase.dshocsinh WHERE MaHS = "S123";'
        #run_cmd(cmd3)
        #for row in recs:
        #    print(row)
        #print("CMD3 Seeking Finished!!")

        #cmd4 = 'DELETE FROM kvdatabase.dshocsinh WHERE MaHS = "S123";'
        #run_cmd(cmd4)
        #print("CMD4 Delete Finished!!")

        #run_cmd(cmd1)
        #for row in recs:
        #    print(row)
        #print("CMD1 update finished!!")

    #Def nút Cập nhật học sinh vào database
    def but_updatestu(self):

        #Setup retrieve data trên table nhập liệu để chèn vào database
        nrow = self.tbwthemhs.rowCount()
        ncol = self.tbwthemhs.columnCount()
        per_record = []
        alldata = []

        for row in range(0,nrow):
            for col in range(0,ncol):
                if self.tbwthemhs.item(row,col) is None:
                    per_record.append('')
                else:
                    item = str(self.tbwthemhs.item(row,col).text())
                    per_record.append(item)
                    
            alldata.append(per_record)
            per_record = []
            
            #cmd2 = 'INSERT INTO kvdatabase.dshocsinh(MaHS,HoTenHS,NamSinh,NgayVaoHoc,Lop,SoDienThoai,TinhTrangHS) VALUES("S123","Zeta",1995,"2020-03-05",9,"0377122966",1);'
            #run_cmd(cmd2)

                       
            #print('Họ tên học sinh: ',alldata[row][0],', kiểu dữ liệu: ',type(alldata[row][0]))
            #print('Năm sinh: ',alldata[row][1],', kiểu dữ liệu: ',type(alldata[row][1]))
            #print('Thời điểm vào học: ',day[row],'/',month[row],'/',year[row],', kiểu dữ liệu: ',type(day[row]),'/',type(month[row]),'/',type(year[row]))
            #print('Số điện thoại: ',alldata[row][3],', kiểu dữ liệu: ',type(alldata[row][3]))
            #print()
            
        #Notification
        self.thongbao.setText("Đã cập nhật học sinh!!")
        
        for row in alldata:
            ipcmd = 'INSERT INTO dshocsinh(HoTenHS,NamSinh,NgayVaoHoc,SoDienThoai,Lop,TinhTrangHS,MaHS) VALUES('
            ipcmd += '"'+row[0]+'",'+str(row[1])+',"'+row[2][4:]+'-'+row[2][2:4]+'-'+row[2][:2]+'","'+row[3]+'",'+str(row[4])+',1,"'+row[5]+'");'

            #Counter start
            startcounter = datetime.datetime.now()
            microsec_start = startcounter.microsecond

            run_cmd(ipcmd)
            print("Command Ran!!")

            #Counter end and output
            endcounter = datetime.datetime.now()
            microsec_end = endcounter.microsecond
            process_time = endcounter - startcounter
            microsec_process = microsec_end - microsec_start
            print('Connected in ',process_time ,' secs and ',microsec_process , ' microseconds!')




# Window 1-3 - Dialog xóa học sinh cũ / nghỉ
class OpWinM13_xoahocsinh(Ui_xoahocsinh):

    #Khởi tạo
    def setupUi(self, window_13):
        super().setupUi(window_13)

        #Button clicked definition
        self.hoantat.clicked.connect(self.but_hoantat)
        self.loc.clicked.connect(self.but_loc)
        self.xoa.clicked.connect(self.but_xoa)
        self.chgstatus.clicked.connect(self.but_chgstatus)

        #Show current defined window
        window_13.show()

    #Def nút Hoàn tất
    def but_hoantat(self):
        window_13.hide()
        window_1.show()

    #Def nút lọc
    def but_loc(self):
        
        #Redef thông báo
        self.thongbao.setText("Chọn thêm học sinh để xóa. Ấn Hoàn tất nếu muốn trở lại trang trước.")

        w13_lop = self.lop.value()

        #Setup các tùy chọn lệnh SQL khi nhập lớp
        if (w13_lop == 0):
            self.cmd = 'SELECT * FROM dshocsinh;'
        elif (w13_lop == 1):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 1;'
        elif (w13_lop == 2):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 2;'
        elif (w13_lop == 3):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 3;'
        elif (w13_lop == 4):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 4;'
        elif (w13_lop == 5):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 5;'
        elif (w13_lop == 6):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 6;'
        elif (w13_lop == 7):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 7;'
        elif (w13_lop == 8):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 8;'
        elif (w13_lop == 9):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 9;'
        elif (w13_lop == 10):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 10;'
        elif (w13_lop == 11):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 11;'
        elif (w13_lop == 12):
            self.cmd = 'SELECT * FROM dshocsinh WHERE Lop = 12;'

        # Xử lý lệnh SQL đã tạo ở trên
        run_cmd(self.cmd)
            
        #Create lines for View table to storage items
        self.tbwviewhs.setRowCount(len(recs))
        self.tbwviewhs.setColumnCount(7)

        self.header1 = self.tbwviewhs.horizontalHeader()       
        self.header1.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.header1.setSectionResizeMode(6, QHeaderView.ResizeToContents)

        for row in range(0,len(recs)):
            for col in range(0,6):
                self.tbwviewhs.setItem(row,col,QTableWidgetItem(str(recs[row][col])))
            if recs[row][6] == b'\x01':
                self.tbwviewhs.setItem(row,6,QTableWidgetItem('Đang học'))
            else:
                self.tbwviewhs.setItem(row,6,QTableWidgetItem('Đã nghỉ'))


        #Def notification hoàn tất việc lọc
        self.locnoti.setText(QtCore.QCoreApplication.translate("xoahocsinh", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600;\">Đã hoàn tất lọc!!!</span></p></body></html>"))
        

    # Def nút xóa
    def but_xoa(self):
        self.del_items = self.mahscanxoa.toPlainText()
        self.del_itlist = []
        self.del_itlist = self.del_items.split(",")
        for mahs in self.del_itlist:
            cmd_delete = 'DELETE FROM dshocsinh WHERE MaHS="' + mahs + '";'
            run_cmd(cmd_delete)
        self.thongbao.setText("")
        self.thongbao.setText("Đã xóa các mã học sinh được chọn!!")


    # Def nút Đổi trạng thái
    def but_chgstatus(self):
        self.ipitems = self.mahsdoitt.toPlainText()
        self.ipidlist = []
        self.ipidlist = self.ipitems.split(',')
        
        self.iprcd = []
        for mahs in self.ipidlist:
            cmd_sortout = 'SELECT * FROM dshocsinh WHERE MaHS = "' + mahs + '";'
            run_cmd(cmd_sortout)
            print(recs)
            self.ouprcd = list(recs[0])
            if  self.ouprcd[6] == b'\x00':
                self.ouprcd[6] = '1'
            else:
                self.ouprcd[6] = '0'
            self.iprcd.append(self.ouprcd)
        print(self.iprcd)

        for row in self.iprcd:
            cmd_update = 'UPDATE dshocsinh SET TinhTrangHS = ' + str(row[6]) + ' WHERE MaHS = "' + row[0] + '";'
            print(cmd_update)
            run_cmd(cmd_update)

#Window 1-4 - Thống kê toàn bộ danh sách điểm danh học sinh
class OpWinM14_thongkediemdanh(Ui_tkdiemdanhhs):

    #Khởi tạo
    def setupUi(self, window_14):
        super().setupUi(window_14)

        #Define default values
        self.datefin.setDate(QtCore.QDate.currentDate())

        #Button clicked definition
        self.hoantat.clicked.connect(self.Op_hoantat)
        self.loc.clicked.connect(self.Op_loc)
        
        #Show current defined window
        window_14.show()


    #Def nút Hoàn tất
    def Op_hoantat(self):
        window_14.hide()
        window_1.show()

    #Def nút lọc
    def Op_loc(self):
        ngaybd = self.datestart.date().toPyDate()
        ngaykt = self.datefin.date().toPyDate()
        ngaybd.strftime('%Y-%m-%d')
        ngaykt.strftime('%Y-%m-%d')
        print(str(ngaybd),' - ',str(ngaykt))
        
        cmd_sl = 'SELECT diemdanhhs.IDDiemDanh,diemdanhhs.MaHS, dshocsinh.HoTenHS, diemdanhhs.MaMonHoc, diemdanhhs.MaNhom, diemdanhhs.NgayGioHoc FROM diemdanhhs INNER JOIN dshocsinh ON diemdanhhs.MaHS = dshocsinh.MaHS WHERE DATE(diemdanhhs.NgayGioHoc) BETWEEN "' + str(ngaybd) + '" AND "' + str(ngaykt) + '";'
        print(cmd_sl)
        run_cmd(cmd_sl)

        #Prepare table for value Input
        self.tbdiemdanh_view.setRowCount(len(recs))
        self.tbdiemdanh_view.setColumnCount(6)

        self.header2 = self.tbdiemdanh_view.horizontalHeader()
        for col in range(0,6):
            self.header2.setSectionResizeMode(col,QHeaderView.ResizeToContents)



        for row in range(0,len(recs)):
            for col in range(0,6):
                self.tbdiemdanh_view.setItem(row,col,QTableWidgetItem(str(recs[row][col])))

        print(cmd_sl)
        for row in recs:
            print(row)

#Window 1-6 - Hiển thị danh sách toàn bộ học sinh
class OpWinM16_danhsachhocsinh(Ui_danhsachhs):

    #Khởi tạo
    def setupUi(self, window_16):
        super().setupUi(window_16)
        #Set Radiobutton check
        self.allstushow.setChecked(True)

        #Button clicked definition
        self.loadview.clicked.connect(self.loaditem)
        self.hoantat.clicked.connect(self.Op_hoantat)
        self.icall.clicked.connect(self.increase_all)
        self.dcall.clicked.connect(self.decrease_all)
        self.icperstu.clicked.connect(self.increase_per)
        self.dcperstu.clicked.connect(self.decrease_per)

        #Show current defined window
        window_16.show()

    #Def nút hoàn tất
    def Op_hoantat(self):
        window_16.hide()
        window_1.show()

    #Def nút load danh sách học sinh
    def loaditem(self):
        signal_get = 1
        if self.curstushow.isChecked():
            signal_get = 0
        #print(signal_get)

        data_cmd = 'SELECT * FROM dshocsinh;'
        if signal_get == 0:
            data_cmd = data_cmd[:-1]
            data_cmd += ' WHERE TinhTrangHS = 1;'

        data_cmd = data_cmd[:-1]
        data_cmd += ' ORDER BY Lop ASC;'

        #print(data_cmd)
        
        run_cmd(data_cmd)
        showdata = []
        for item in recs:
            peritem = []
            for it in item:
                peritem.append(it)
            tinhtrang = str(peritem[6])
            #print(tinhtrang[5])
            if tinhtrang[5] == '0':
                peritem[6] = 1
            else:
                peritem[6] = 0
                

            showdata.append(peritem)
            
            #print(item)
        #for item in showdata:
        #    print(item)

        #Update dữ liệu lên view
        self.dataview.setRowCount(len(showdata))
        self.dataview.setColumnCount(7)
        self.header21 = self.dataview.horizontalHeader()
        for it in range(0,7):
            self.header21.setSectionResizeMode(it,QHeaderView.ResizeToContents)

        for row in range(0,len(showdata)):
            for col in range(0,7):
                self.dataview.setItem(row,col,QTableWidgetItem(str(showdata[row][col])))

                


    def increase_all(self):
        cmd_remove_safe_updates = 'SET SQL_SAFE_UPDATES = 0;'
        cmd_recall_safe_updates = 'SET SQL_SAFE_UPDATES = 1;'
        cmd_icall = 'UPDATE dshocsinh SET Lop = Lop + 1;'
        
        run_cmd(cmd_remove_safe_updates)
        run_cmd(cmd_icall)
        run_cmd(cmd_recall_safe_updates)

    def decrease_all(self):
        cmd_remove_safe_updates = 'SET SQL_SAFE_UPDATES = 0;'
        cmd_recall_safe_updates = 'SET SQL_SAFE_UPDATES = 1;'
        cmd_icall = 'UPDATE dshocsinh SET Lop = Lop - 1;'
        
        run_cmd(cmd_remove_safe_updates)
        run_cmd(cmd_icall)
        run_cmd(cmd_recall_safe_updates)

    def increase_per(self):
        ipstucode = self.mahsicdc.text()
        print(ipstucode)

        cmd_remove_safe_updates = 'SET SQL_SAFE_UPDATES = 0;'
        cmd_recall_safe_updates = 'SET SQL_SAFE_UPDATES = 1;'
        cmd_icall = 'UPDATE dshocsinh SET Lop = Lop + 1 WHERE MaHS = "' + ipstucode + '";'
        
        run_cmd(cmd_remove_safe_updates)
        run_cmd(cmd_icall)
        run_cmd(cmd_recall_safe_updates)




    def decrease_per(self):
        ipstucode = self.mahsicdc.text()
        print(ipstucode)

        cmd_remove_safe_updates = 'SET SQL_SAFE_UPDATES = 0;'
        cmd_recall_safe_updates = 'SET SQL_SAFE_UPDATES = 1;'
        cmd_icall = 'UPDATE dshocsinh SET Lop = Lop - 1 WHERE MaHS = "' + ipstucode + '";'
        
        run_cmd(cmd_remove_safe_updates)
        run_cmd(cmd_icall)
        run_cmd(cmd_recall_safe_updates)














#This part intended left blank




#Window 2 - Dialog Quản lý giáo viên
class OpWinM2_qlgv(Ui_quanlygiaovien):
    #Khởi tạo
    def setupUi(self,window_2):
        super().setupUi(window_2)
        
        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.dsgv.clicked.connect(self.tbslgvct)
        self.themgiaovien.clicked.connect(self.themgvbut)
        self.cnpcmn.clicked.connect(self.pcgv)
        self.xoagv.clicked.connect(self.xgv)

        #Show current defined window
        window_2.show()

    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_2.hide()

    #Def nút danh sách giáo viên
    def tbslgvct(self):
        self.ui = OpWinM21_slgvct()
        self.ui.setupUi(window_21)
        window_2.hide()

    #Def nút thêm giáo viên
    def themgvbut(self):
        self.ui = OpWinM22_themgv()
        self.ui.setupUi(window_22)
        window_2.hide()

    #Def nút phân công giáo viên
    def pcgv(self):
        self.ui = OpWinM23_pcgv()
        self.ui.setupUi(window_23)
        window_2.hide()

    #Def nút xóa giáo viên
    def xgv(self):
        self.ui = OpWin24_xoagv()
        self.ui.setupUi(window_24)
        window_2.hide()





#Open Window 2-1 - Hiển thị danh sách giáo viên của cơ sở
class OpWinM21_slgvct(Ui_dsgv):

    #Init def
    def __init__(self):
        cmd_dsmon = 'SELECT MaMonHoc FROM dmmonhoc;'
        run_cmd(cmd_dsmon)
        self.dsmon = ['0']
        for item in recs:
            self.dsmon.append(item[0])


    #Khởi tạo
    def setupUi(self,window_21):
        super().setupUi(window_21)

        #Init danh sách môn học
        self.monday.addItems(self.dsmon)

        #Button clicked definition
        self.hoantat.clicked.connect(self.ht)
        self.loc.clicked.connect(self.butloc)

        #Show current window
        window_21.show()


    #def nút hoàn tất
    def ht(self):
        window_2.show()
        window_21.hide()

    #def nút lọc giáo viên
    def butloc(self):
        mamon = self.monday.currentText()
        if mamon != '0':
            cmd_laymon = 'SELECT dsgiaovien.MaGV, dsgiaovien.tenGV FROM dsgiaovien INNER JOIN phanconggiangday ON phanconggiangday.maGV = dsgiaovien.MaGV WHERE phanconggiangday.MaMonHoc = "' + mamon + '";'
            print(cmd_laymon)
            run_cmd(cmd_laymon)
            #print(recs)
            allitem = []
            for row in recs:
                peritem = []
                peritem.append(row[0])
                peritem.append(row[1])
                allitem.append(peritem)
            print(allitem)

        else:
            cmd_laymon = 'SELECT MaGV, tenGV FROM dsgiaovien;'
            print(cmd_laymon)
            run_cmd(cmd_laymon)
            #print(recs)
            allitem = []
            for row in recs:
                peritem = []
                peritem.append(row[0])
                peritem.append(row[1])
                allitem.append(peritem)
            print(allitem)
            

        #Push lên list view
        self.gv_view.setRowCount(len(allitem))
        self.gv_view.setColumnCount(2)
        self.header14 = self.gv_view.horizontalHeader()
        for col in range(0,2):
            self.header14.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        for row in range(0,len(allitem)):
            self.gv_view.setItem(row,0,QTableWidgetItem(allitem[row][0]))
            self.gv_view.setItem(row,1,QTableWidgetItem(allitem[row][1]))
            


#Open Window 2-2 - Thêm giáo viên mới vào cơ sở dữ liệu
class OpWinM22_themgv(Ui_themgv):

    #Khởi tạo
    def setupUi(self, window_22):
        super().setupUi(window_22)

        #Button click definition
        self.hoantat.clicked.connect(self.ql)
        self.but_themgv.clicked.connect(self.thtgvmoi)
        self.updatedata.clicked.connect(self.uddt)



        #show current window
        window_22.show()

    #Def nút hoàn tất
    def ql(self):
        window_2.show()
        window_22.hide()

    #Def nút tiến hành thêm giáo viên mới
    def thtgvmoi(self):
        sogv = self.sogvmoi.value()
        print(sogv)

        #Lấy max id
        cmd_idgv = 'SELECT maGV FROM dsgiaovien;'
        run_cmd(cmd_idgv)
        idgvlist = []
        for row in recs:
            idgvlist.append(row[0])
        
        for i in range(0,len(idgvlist)):
            idgvlist[i] = idgvlist[i][1:]
            if idgvlist[i].isdigit():
                idgvlist[i] = int(idgvlist[i])
            else:
                idgvlist[i] = 0
        print(idgvlist)
        maxid = max(idgvlist)

        #Tạo bảng hiển thị
        self.addgv_view.setRowCount(sogv)
        self.addgv_view.setColumnCount(2)
        self.header15 = self.addgv_view.horizontalHeader()

        for col in range(0,2):
            self.header15.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        listnewid = []
        for i in range(0,sogv):
            maxid += 1
            ipid = 'T' + str(maxid)
            listnewid.append(ipid)
        print(listnewid)


        #Chèn lên bảng hiển thị
        for row in range(0,sogv):
            self.addgv_view.setItem(row,0,QTableWidgetItem(listnewid[row]))

    #Def cập nhật dữ liệu mới vào cơ sở dữ liệu
    def uddt(self):
        numgv = self.sogvmoi.value()

        alldata = []
        for row in range(0,numgv):
            it0 = str(self.addgv_view.item(row,0).text())
            it1 = str(self.addgv_view.item(row,1).text())
            perrow = []
            perrow.append(it0)
            perrow.append(it1)
            alldata.append(perrow)

        print(alldata)

        #Cập nhật vào cơ sở dữ liệu
        for row in range(0,len(alldata)):
            cmd_udata = 'INSERT INTO dsgiaovien (maGV,tenGV,PayRate) VALUES ("' + alldata[row][0] + '", "' + alldata[row][1] + '", "0");'
            print(cmd_udata)
            run_cmd(cmd_udata)

#Open Window 2-3 - Phân công giáo viên
class OpWinM23_pcgv(Ui_M23_capnhatxoapc):
    
    #def init
    def __init__(self):
        cmd_locmagv = 'SELECT maGV FROM dsgiaovien;'
        run_cmd(cmd_locmagv)
        self.dsmagiaovien = []
        for row in recs:
            self.dsmagiaovien.append(row[0])

        cmd_listid = 'SELECT IDPhanCong FROM phanconggiangday;'
        run_cmd(cmd_listid)
        self.idlist = []
        for item in recs:
            self.idlist.append(str(item[0]))

    #Khởi tạo
    def setupUi(self,window_23):
        super().setupUi(window_23)

        #Khởi động với việc tải bảng 
        self.updatetable()
        #Combobox mã giáo viên
        self.magvadd.addItems(self.dsmagiaovien)
        self.magvadd.currentIndexChanged.connect(self.onupdate_mgvadd)
        
        #Combobox ID Phân công
        self.mapccapnhat.addItems(self.idlist)
        self.mapccapnhat.currentIndexChanged.connect(self.onupdate_mapccapnhat)

        #Tắt hiển thị các items khởi tạo
        self.mamonpc.setEnabled(False)
        self.mamondoi.setEnabled(False)
        

        #Button click definition
        self.hoantat.clicked.connect(self.ql)
        self.thempc.clicked.connect(self.tpc)
        self.updatepc.clicked.connect(self.cnpc)
        self.xoapc.clicked.connect(self.xoaidpc)


        #Show window
        window_23.show()

    #Def nút quay lại
    def ql(self):
        window_2.show()
        window_23.hide()

    #Def khi cập nhật mã phân công
    def onupdate_mapccapnhat(self):
        self.updatetable()
        self.mamondoi.setEnabled(True)
        cmd_info = 'SELECT maGV FROM phanconggiangday WHERE IDPhanCong = ' + self.mapccapnhat.currentText() + ';'
        print(cmd_info)
        run_cmd(cmd_info)
        #print(recs)
        mgv_chg = recs[0][0]
        cmd_allmm = 'SELECT MaMonHoc FROM dmmonhoc;'
        run_cmd(cmd_allmm)
        allmm = []
        for row in recs:
            allmm.append(row[0])

        cmd_mmgv = 'SELECT MaMonHoc FROM phanconggiangday WHERE maGV = "' + mgv_chg + '";'
        run_cmd(cmd_mmgv)
        mmgv = []
        for row in recs:
            mmgv.append(row[0])

        mmshow = []
        for row in allmm:
            if row not in mmgv:
                mmshow.append(row)

        print(mmshow)
        
        self.mamondoi.clear()
        self.mamondoi.addItems(mmshow)

    #Cập nhật bảng hiển thị
    def updatetable(self):
        cmdupdate = 'SELECT phanconggiangday.IDPhanCong, phanconggiangday.maGV, dsgiaovien.tenGV, phanconggiangday.MaMonHoc, dmmonhoc.TenMonHoc FROM phanconggiangday INNER JOIN dsgiaovien ON phanconggiangday.maGV = dsgiaovien.maGV INNER JOIN dmmonhoc ON phanconggiangday.MaMonHoc = dmmonhoc.MaMonHoc;'
        print(cmdupdate)
        run_cmd(cmdupdate)
        print(recs)

        #Tạo bảng và đẩy dữ liệu vào bảng
        self.dscnpc_view.setColumnCount(5)
        self.dscnpc_view.setRowCount(len(recs))
        self.header16 = self.dscnpc_view.horizontalHeader()
        for col in range(0,5):
            self.header16.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Load data lên tablewidget
        for row in range(0,len(recs)):
            for col in range(0,5):
                self.dscnpc_view.setItem(row,col,QTableWidgetItem(str(recs[row][col])))

    #Def khi cập nhật mã môn
    def onupdate_mgvadd(self):
        cmdloc = 'SELECT MaMonHoc FROM dmmonhoc;'
        run_cmd(cmdloc)
        allmm = []
        for row in recs:
            allmm.append(row[0])
        print(allmm)


        cmdlocmmgv = 'SELECT MaMonHoc FROM phanconggiangday WHERE maGV = "' + self.magvadd.currentText() + '";'
        run_cmd(cmdlocmmgv)
        #print(recs)
        teamm = []
        for row in recs:
            if row[0] not in teamm:
                teamm.append(row[0])
        print(teamm)

        ipmmpc = []
        for item in allmm:
            if item not in teamm:
                ipmmpc.append(item)
        print(ipmmpc)

        self.mamonpc.setEnabled(True)
        self.mamonpc.clear()
        self.mamonpc.addItems(ipmmpc)
        self.updatetable()

    #Def nút thêm phân công
    def tpc(self):
        mgv = self.magvadd.currentText()
        mmpc = self.mamonpc.currentText()
        
        #Lấy max id phân công
        cmdmax ='SELECT MAX(IDPhanCong) FROM phanconggiangday;'
        run_cmd(cmdmax)
        maxid = recs[0][0]
        print(maxid)

        #Cập nhật info vào bảng phân công giáo viên
        idip = maxid + 1
        cmdip = 'INSERT INTO phanconggiangday VALUES (' + str(idip) + ', "' + mgv + '", "' + mmpc + '", "' + mgv + '", "' + mmpc + '");'
        print(cmdip)
        run_cmd(cmdip)

        #Load lại bảng view
        self.updatetable()
        self.onupdate_mgvadd()


    #Def nút cập nhật phân công
    def cnpc(self):
        cmd_update = 'UPDATE phanconggiangday SET MaMonHoc = "' + self.mamondoi.currentText() + '", dmmonhoc_MaMonHoc = "' + self.mamondoi.currentText() + '" WHERE IDPhanCong = ' + self.mapccapnhat.currentText() +';'
        print(cmd_update)
        run_cmd(cmd_update)
        self.updatetable()
        self.onupdate_mapccapnhat()


    #Def nút xóa id phân công
    def xoaidpc(self):
        
        idstr = self.mapcdexoa.text()
        ipids = []
        print(idstr)
        ipids = idstr.split(',')
        print(ipids)

        #Tiến hành xóa
        for item in ipids:
            cmdxoa = 'DELETE FROM phanconggiangday WHERE IDPhanCong = ' + item + ';'
            run_cmd(cmdxoa)

        self.updatetable()


#Open Window 2-4 - Xóa giáo viên
class OpWin24_xoagv(Ui_formxoagv):

    #On Init
    def __init__(self):

        #Lấy danh sách giáo viên từ bảng dsgiaovien
        cmd_slmagv = 'SELECT maGV FROM dsgiaovien;'
        run_cmd(cmd_slmagv)
        self.ip_magv = []
        for row in recs:
            #print(row)
            self.ip_magv.append(row[0])

        

    #Khởi tạo
    def setupUi(self,window_24):
        super().setupUi(window_24)

        #Đẩy dữ liệu lên combobox
        self.magv_ip.addItems(self.ip_magv)

        #Khi cập nhật tên giáo viên trên lineedit
        self.magv_ip.activated.connect(self.update_hotengv)

        #Khi cập nhật tên giáo viên đưa ra khỏi danh sách xóa khi line mã giáo viên để xóa thay đổi
        self.magv_del.textChanged.connect(self.delcb)

        #Khi cập nhật mã giáo viên để loại ra khỏi danh sách xóa
        self.magv_out.activated.connect(self.update_hotenout)


        #Button click definition
        self.hoantat.clicked.connect(self.ql)
        self.nhapdexoa.clicked.connect(self.update_dellist)
        self.loaikhoids.clicked.connect(self.getoutgv)
        self.xoagv.clicked.connect(self.xgv)

        #Show window
        window_24.show()

    #Def nút quay lại
    def ql(self):
        window_2.show()
        window_24.hide()

    #Def cập nhật line họ tên giáo viên
    def update_hotengv(self):
        cmd_hoten = 'SELECT tenGV FROM dsgiaovien WHERE maGV = "' + self.magv_ip.currentText() + '";'
        #print(cmd_hoten)
        run_cmd(cmd_hoten)
        #print(recs)
        ip_ht = recs[0][0]
        self.hotengv.setText('')
        self.hotengv.setText(ip_ht)


    #Def but nhập vào danh sách xóa
    def update_dellist(self):

        #Tắt hiển thị danh sách xóa
        self.magv_del.setEnabled(False)

        ipmgv = self.magv_ip.currentText()
        current_dellist_str = self.magv_del.text()
        if len(current_dellist_str) == 0:
            current_dellist_str += ipmgv + ','
        else:
            check_list = []
            check_list = current_dellist_str.split(',')
            if ipmgv not in check_list:
                current_dellist_str += ipmgv + ','

        self.magv_del.setText(current_dellist_str)

        print(current_dellist_str)

    #Def khi cập nhật thêm danh sách giáo viên để xóa
    def delcb(self):
        ip_magv = self.magv_del.text()
        ip_magv = ip_magv[:len(ip_magv)-1]
        #print(ip_magv)
        oplistmagv = ip_magv.split(',')
        #print(oplistmagv)
        
        self.magv_out.clear()
        self.magv_out.addItems(oplistmagv)
        


    #Def khi cập nhật combobox đưa giáo viên ra khỏi danh sách xóa
    def update_hotenout(self):
        cmd_ht = 'SELECT tenGV FROM dsgiaovien WHERE maGV = "' + self.magv_out.currentText() + '";'
        run_cmd(cmd_ht)
        self.hotengvout.setText(recs[0][0])
        #print(cmd_ht)
        #print(recs[0][0])

    #Def khi nhấn nút đưa giáo viên ra khỏi danh sách xóa
    def getoutgv(self):
        getitout = self.magv_out.currentText()

        ipstr = self.magv_del.text()
        ipstr = ipstr[:len(ipstr)-1]
        ip_dellist = ipstr.split(',')
        
        opstr = ''
        for item in ip_dellist:
            if item != getitout:
                opstr += item + ','

        #print(opstr)
        #print(ip_dellist)
        self.magv_del.setText(opstr)


    #Def nút xóa giáo viên trên danh sách xóa
    def xgv(self):
        strxoa = self.magv_del.text()
        strxoa = strxoa[:len(strxoa)-1]
        listxoa = strxoa.split(',')
        #print(listxoa)

        for item in listxoa:
            cmd_xoa = 'DELETE FROM dsgiaovien WHERE maGV = "' + item + '";'
            run_cmd(cmd_xoa)

        #Cập nhật lại các list add giáo viên để xóa trong current dialog box
        cmd_udmagv = 'SELECT maGV FROM dsgiaovien;'
        run_cmd(cmd_udmagv)
        print(recs)

        mgv_list = []
        for record in recs:
            mgv_list.append(record[0])

        print(mgv_list)
        

        
        #Cập nhật lại dialog box
        self.magv_ip.clear()
        self.magv_ip.addItems(mgv_list)
        
        self.magv_out.clear()

        self.magv_del.setText('')









#This part intended left blank




#Window 3 - Dialog Quản lý học phí
class OpWinM3_qlhp(Ui_quanlyhocphi):

    #Khởi tạo
    def setupUi(self,window_3):
        super().setupUi(window_3)

        # Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.xhpdmc.clicked.connect(self.Op_xoasuahpdm)
        self.themhpdm.clicked.connect(self.Op_thpdm)
        self.gndth.clicked.connect(self.Op_gnth)

        #Show current defined window
        window_3.show()
    

    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_3.hide()

    #Def nút mở window_31
    def Op_xoasuahpdm(self):
        self.ui = OpWinM31_qlhp()
        self.ui.setupUi(window_31)
        window_3.hide()

    #Def nút mở window_32
    def Op_thpdm(self):
        self.ui = OpWinM32_thpdm()
        self.ui.setupUi(window_32)
        window_3.hide()

    #Def nút mở window_33
    def Op_gnth(self):
        self.ui = OpWinM33_gnthdshp()
        self.ui.setupUi(window_33)
        window_3.hide()


#Window 3-1 - Dialog Xóa, sửa, thêm định mức học phí
class OpWinM31_qlhp(Ui_xoahpdm):

    #Khởi tạo
    def setupUi(self, window_31):
        super().setupUi(window_31)
        
        #Tắt hiển thị nút cập nhật
        self.capnhat.setEnabled(False)

        #Button clicked definition
        self.hoantat.clicked.connect(self.finished)
        self.loc.clicked.connect(self.Op_loc)
        self.capnhat.clicked.connect(self.Op_capnhat)
        self.delidhp.clicked.connect(self.Op_xoaid)

        #Signal emit when qtablewidget changed
        self.defhp_view.itemChanged.connect(self.Op_enbcapnhat)

        #Show current defined window
        window_31.show()

    #Def khi cập nhật bảng view
    def Op_enbcapnhat(self):
        self.capnhat.setEnabled(True)


    #Def nút quay lại
    def finished(self):
        window_31.hide()
        window_3.show()

    #Def nút chuyển số integer thành chuỗi có dot seperate
    def viewnum(self,ipnum):
        strnum = str(ipnum)
        #print(strnum)
        liststr = []
        while len(strnum)>0:
            if len(strnum) >= 3:
                #print(strnum[-4:-1])
                liststr.append(strnum[len(strnum)-3:])
                strnum = strnum[:len(strnum)-3]
                #print(strnum)
                
            else:
                liststr.append(strnum)
                strnum = ''

        opstr = ''
        for item in liststr:
          opstr = item + '.' + opstr
        
        opstr = opstr[:-1]

        return opstr


    def Op_loc(self):
        self.info_before_event = list()

        #CMD lọc từ bảng defhocphi
        if self.lop.value() == 0:
            cmd_sort = 'SELECT defhocphi.IDhocphi, defhocphi.MaHS, dshocsinh.HoTenHS, defhocphi.MaMonHoc, defhocphi.HocPhi, defhocphi.MaGV, defhocphi.SoBuoiMoiLanDongTien, defhocphi.SoLanDaDongTien, defhocphi.SoBuoiDaHoc FROM defhocphi INNER JOIN dshocsinh ON dshocsinh.MaHS = defhocphi.MaHS;'
        else:
            cmd_sort = 'SELECT defhocphi.IDhocphi, defhocphi.MaHS, dshocsinh.HoTenHS, defhocphi.MaMonHoc, defhocphi.HocPhi, defhocphi.MaGV, defhocphi.SoBuoiMoiLanDongTien, defhocphi.SoLanDaDongTien, defhocphi.SoBuoiDaHoc FROM defhocphi INNER JOIN dshocsinh ON dshocsinh.MaHS = defhocphi.MaHS WHERE dshocsinh.Lop = ' + str(self.lop.value()) + ';'
        
        run_cmd(cmd_sort)
        
        #Sao lưu dữ liệu để đối chiếu khi hoàn tất
        for row in recs:
            self.info_before_event.append(row)

        #Chuẩn bị tableview để xuất dữ liệu
        self.defhp_view.setRowCount(len(recs))
        self.defhp_view.setColumnCount(10)

        self.header3 = self.defhp_view.horizontalHeader()
        for col in range(0,10):
            self.header3.setSectionResizeMode(col,QHeaderView.ResizeToContents)
       

        #Nhập liệu vào bảng view
        for row in range(len(recs)):
            for col in range(0,4):
                self.defhp_view.setItem(row,col,QTableWidgetItem(str(recs[row][col])))
            
            self.defhp_view.setItem(row,4,QTableWidgetItem(str(self.viewnum(recs[row][4]))))

            for col in range(5,9):
                self.defhp_view.setItem(row,col,QTableWidgetItem(str(recs[row][col])))

            if (recs[row][6] == 0):
                ipCurStat = 0
                sbchuadong = 0
            else:
                ipCurStat = str(Fraction(int(recs[row][8])/int(recs[row][6]) - int(recs[row][7])))
                sbchuadong = recs[row][8] - (recs[row][6] * recs[row][7])
                ipCurStat = ipCurStat + ' ----> ' + str(sbchuadong)
            
            

            #print(ipCurStat)
            self.defhp_view.setItem(row,9,QTableWidgetItem(ipCurStat))

            #Set màu cho item
            if (recs[row][6] == 0):
                colorval = 0
            else:
                colorval = int(recs[row][8])/int(recs[row][6]) - int(recs[row][7])

            if 0.75 <= colorval < 1:
                self.defhp_view.item(row,9).setBackground(QtGui.QColor(118,73,243,127))
            elif colorval >=1:
                self.defhp_view.item(row,9).setBackground(QtGui.QColor(255,0,0,127))
            elif 0 <= colorval <= 0.75:
                self.defhp_view.item(row,9).setBackground(QtGui.QColor(3,206,118,127))
            elif colorval < 0:
                self.defhp_view.item(row,9).setBackground(QtGui.QColor(241,249,200,127))


    #Def button cập nhật
    def Op_capnhat(self):
        for row in self.info_before_event:
            print(row)
        self.info_after_event = list()
        for row in range(0,self.defhp_view.rowCount()):
            per_row = list()
            #print(self.defhp_view.columnCount())
            item = int(self.defhp_view.item(row,0).text())
            per_row.append(item)
            
            for col in range(1,4):
                item = str(self.defhp_view.item(row,col).text())
                per_row.append(item)

            item = int(self.defhp_view.item(row,4).text().replace('.',''))
            per_row.append(item)
            
            item = str(self.defhp_view.item(row,5).text())
            per_row.append(item)
            
            for col in range(6,self.defhp_view.columnCount()-1):
                item = int(self.defhp_view.item(row,col).text())
                per_row.append(item)
            item = int()

            self.info_after_event.append(per_row)

        #Kiểm tra dữ liệu sau chỉnh sửa
        for row in self.info_after_event:
            print(row)

        #Kiểm tra độ tương thích giữa 2 list trước và sau
        list_ck = list()
        for cs in range(0,len(self.info_before_event)):
            t = 0
            for row in range(0,len(self.info_before_event[cs])):
                if str(self.info_before_event[cs][row]) != str(self.info_after_event[cs][row]):
                    t = 1
            list_ck.append(t)
        print(list_ck)
        
        cmd_call = ''
        for ck in range(0,len(list_ck)):
            if list_ck[ck] == 1:
                for col in range(0,len(self.info_before_event[ck])):
                    if str(self.info_after_event[ck][col]) != str(self.info_before_event[ck][col]):
                        cmd_call = 'UPDATE defhocphi SET '
                        #print(self.info_after_event[ck][col], ' - ',self.info_before_event[ck][col])
                        if col in (3,5):
                            if col == 3:
                                cmd_call += 'MaMonHoc= "' + str(self.info_after_event[ck][col]) + '" WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                            elif col == 5:
                                cmd_call += 'MaGV= "' + str(self.info_after_event[ck][col]) + '" WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                        if col in (4,6,7,8):
                            if col == 4:
                                cmd_call += 'HocPhi= ' + str(self.info_after_event[ck][col]).replace('.','') + ' WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                            elif col == 6:
                                cmd_call += 'SoBuoiMoiLanDongTien= ' + str(self.info_after_event[ck][col]) + ' WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                            elif col == 7:
                                cmd_call += 'SoLanDaDongTien= ' + str(self.info_after_event[ck][col]) + ' WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                            elif col == 8:
                                cmd_call += 'SoBuoiDaHoc= ' + str(self.info_after_event[ck][col]) + ' WHERE IDhocphi= ' + str(self.info_after_event[ck][0]) +';'
                    if cmd_call != '':
                        print(col)
                        print(cmd_call)
                        #Gọi lệnh cập nhật dữ liệu
                        run_cmd(cmd_call)
                        cmd_call = ''

    def Op_xoaid(self):
        print("Start delete ID!!!")
        self.del_idhpdm = self.lineEdit.text()
        self.del_idhpdmlist = []
        self.del_idhpdmlist_str = self.del_idhpdm.split(',')
        self.del_idhpdm_int = []
        for item in self.del_idhpdmlist_str:
            self.del_idhpdm_int.append(int(item))
        print(self.del_idhpdm_int)

        #Tiến hành tạo và chạy lệnh xóa
        for row in self.del_idhpdm_int:
            cmd_del = 'DELETE FROM defhocphi WHERE IDhocphi = ' + str(row) + ';'
            print(cmd_del)

            #Chạy lệnh xóa
            run_cmd(cmd_del)

        #Xác nhận delete
        print('IDs Deleted!!')








#Window 3-2 - Thêm học phí định mức
class OpWinM32_thpdm(Ui_themidhp):

    #Khởi tạo
    def setupUi(self, window_32):
        super().setupUi(window_32)

        #Button clicked definition
        self.hoantat.clicked.connect(self.finished)
        self.xacnhan.clicked.connect(self.but_xacnhan)
        self.updatetableinfo.clicked.connect(self.Op_updatetable)
        self.updateintodb.clicked.connect(self.Op_updateintodb)

        #Hiển thị window 32
        window_32.show()


    #Def finished button
    def finished(self):
        window_32.hide()
        window_3.show()

    #Def button xác nhận
    def but_xacnhan(self):
        #Xóa thông tin cũ trong trường hợp dùng lại widget ngay
        while (self.themidhp_view.rowCount()>0):
            self.themidhp_view.removeRow(0)

        #Lấy max IDhocphi và lấy số học sinh cần thêm
        self.ipsohs = self.sohs.value()
        cmd_max = 'SELECT MAX(IDhocphi) FROM defhocphi;'
        run_cmd(cmd_max)
        maxid = recs[0][0]
        
        #Test
        print(maxid)
        print(type(maxid))
        print('Số học sinh cần thêm: ',self.ipsohs)


        #Chuẩn bị table để nhập dữ liệu mới
        self.themidhp_view.setRowCount(self.ipsohs)
        self.themidhp_view.setColumnCount(8)
        self.header4 = self.themidhp_view.horizontalHeader()
        for col in range(0,8):
            self.header4.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        

        #Xử lý chèn IDhocphi mới
        counter = maxid
        for row in range(0,self.ipsohs):
            counter += 1
            self.themidhp_view.setItem(row,0,QTableWidgetItem(str(counter)))

        #Xử lý chọn mã học sinh mới (Mã này phải thêm trước từ phần quản lý học sinh)
        
        #Lấy data mã học sinh để lựa chọn
        cmd_mahs = 'SELECT MaHS FROM dshocsinh;'
        run_cmd(cmd_mahs)
        dsmahs = []
        #print(len(recs))
        for row in range(0,len(recs)):
            print(recs[row][0])
            dsmahs.append(recs[row][0])
        #print(dsmahs)


        #Đưa lên bảng view
        for row in range(0,self.ipsohs):
            #Dồn data vào list
            combo_mahs = QComboBox()
            for item in dsmahs:
                combo_mahs.addItem(item)
            self.themidhp_view.setCellWidget(row,1,combo_mahs)

        
        #Lấy data giáo viên để lựa chọn
        cmd_magv = 'SELECT MaGV FROM dsgiaovien;'
        run_cmd(cmd_magv)
        dsgv = []
        for row in range(0,len(recs)):
            dsgv.append(recs[row][0])

        #Đưa lên bảng view
        for row in range(0,self.ipsohs):
            #Dồn data vào list
            combo_magv = QComboBox()
            for item in dsgv:
                combo_magv.addItem(item)
            self.themidhp_view.setCellWidget(row,6,combo_magv)


        #Lấy data môn học để lựa chọn
        cmd_mamh = 'SELECT MaMonHoc FROM dmmonhoc;'
        run_cmd(cmd_mamh)
        dsmh = []
        for row in range(0,len(recs)):
            dsmh.append(recs[row][0])

        #Đưa lên bảng view
        for row in range(0,self.ipsohs):
            #Dồn data vào list
            combo_mamh = QComboBox()
            for item in dsmh:
                combo_mamh.addItem(item)
            self.themidhp_view.setCellWidget(row,3,combo_mamh)




    #Def button cập nhật họ tên học sinh, họ tên giáo viên và tên môn học
    def Op_updatetable(self):
        #Các biến lấy thông tin Họ tên học sinh, Môn học, Họ tên giáo viên
        stuname_lt = []
        teaname_lt = []
        subjname_lt = []

        for row in range(0,self.ipsohs):
            #Update họ tên học sinh
            cmd_stuname = 'SELECT HoTenHS FROM dshocsinh WHERE MaHS = "'
            cmd_teaname = 'SELECT tenGV FROM dsgiaovien WHERE maGV = "'
            cmd_subjname = 'SELECT TenMonHoc FROM dmmonhoc WHERE MaMonHoc = "'

            item_mahs = self.themidhp_view.cellWidget(row, 1)
            item_magv = self.themidhp_view.cellWidget(row, 6)
            item_mmh = self.themidhp_view.cellWidget(row, 3)

            if isinstance(item_mahs,QComboBox):
                cur_mahs = item_mahs.currentText()
            if isinstance(item_magv,QComboBox):
                cur_magv = item_magv.currentText()
            if isinstance(item_mmh,QComboBox):
                cur_mmh = item_mmh.currentText()


            cmd_stuname += str(cur_mahs) +'";'
            run_cmd(cmd_stuname)
            stuname_lt.append(recs[0][0])

            cmd_teaname += str(cur_magv) +'";'
            run_cmd(cmd_teaname)
            teaname_lt.append(recs[0][0])

            cmd_subjname += str(cur_mmh) + '";'
            run_cmd(cmd_subjname)
            subjname_lt.append(recs[0][0])


        print(stuname_lt)
        print(teaname_lt)
        print(subjname_lt)

        for row in range(0,self.ipsohs):
            self.themidhp_view.setItem(row,2,QTableWidgetItem(stuname_lt[row]))
            self.themidhp_view.setItem(row,4,QTableWidgetItem(subjname_lt[row]))
            self.themidhp_view.setItem(row,7,QTableWidgetItem(teaname_lt[row]))


    #Def nút cập nhật vào cơ sở dữ liệu
    def Op_updateintodb(self):
        alldata = []
        for row in range(0,self.ipsohs):
            per_row = []

            #Lấy cột 1 - Mã defhocphi
            item_1 = str(self.themidhp_view.item(row,0).text())
            per_row.append(item_1)

            #Lấy cột 2 - Mã học sinh
            item_2 = self.themidhp_view.cellWidget(row,1)
            if isinstance(item_2,QComboBox):
                cur_mahs = item_2.currentText()
            per_row.append(cur_mahs)

            #Lấy cột 3 - Mã môn học
            item_3 = self.themidhp_view.cellWidget(row,3)
            if isinstance(item_3,QComboBox):
                cur_mmh = item_3.currentText()
            per_row.append(cur_mmh)

            #Lấy cột 4 - Học phí
            item_4 = str(self.themidhp_view.item(row,5).text())
            per_row.append(item_4)

            #Lấy cột 5 - Mã giáo viên
            item_5 = self.themidhp_view.cellWidget(row,6)
            if isinstance(item_5,QComboBox):
                cur_magv = item_5.currentText()
            per_row.append(cur_magv)

            #Lưu vào alldata
            alldata.append(per_row)

        #Tiến hành tạo câu lệnh cập nhật
        for row in range(0,len(alldata)):
            cmd_ipdata = 'INSERT INTO defhocphi (IDhocphi,MaHS,MaMonHoc,HocPhi,MaGV,SoBuoiMoiLanDongTien,SoLanDaDongTien,SoBuoiDaHoc,dmmonhoc_MaMonHoc,dshocsinh_MaHS,dsgiaovien_maGV) VALUES (' + alldata[row][0] + ', "' + alldata[row][1] + '", "' + alldata[row][2] + '", '+ alldata[row][3] + ', "' + alldata[row][4] + '", 0, 0, 0, "' + alldata[row][2] + '", "' + alldata[row][1] + '", "' + alldata[row][4] +'");'
            print(cmd_ipdata)
            run_cmd(cmd_ipdata)
        #print(alldata)







#Window 3 - 3 - Ghi nhận tiền học & danh sách học phí đã đóng trong ngày / tháng / năm
class OpWinM33_gnthdshp(Ui_ghinhanvadshp):

    #khởi tạo
    def setupUi(self, window_33):
        super().setupUi(window_33)

        #Button click definition
        self.hoantat.clicked.connect(self.finished)
        self.loc.clicked.connect(self.but_loc)
        self.saveinfo.clicked.connect(self.saveidhp)
        self.delid.clicked.connect(self.xoaidnham)


        #Define startup figures
        self.dateinfo.setDate(QtCore.QDate.currentDate())
        self.daysort.setChecked(True)

        #Show current defined window
        window_33.show()

    #Def nút hoàn tất
    def finished(self):
        window_33.hide()
        window_3.show()

    #Def nút lọc
    def but_loc(self):
        #Lấy options lọc
        stat_ip = 1
        if self.monthsort.isChecked():
            stat_ip = 2
        elif self.yearsort.isChecked():
            stat_ip = 3
        print(stat_ip)

        #Lấy thông tin để lọc
        date_taken = self.dateinfo.date().toPyDate()
        strdate = str(date_taken)
        year,month,date = strdate.split('-')

        #Tạo command lọc 
        cmd_sort = 'SELECT thuhocphi.IDThuHP, thuhocphi.IDhocphi, thuhocphi.NgayThuHP, thuhocphi.MaHS, dshocsinh.HoTenHS, dmmonhoc.TenMonHoc, dsgiaovien.tenGV FROM thuhocphi INNER JOIN dshocsinh ON dshocsinh.MaHS = thuhocphi.MaHS INNER JOIN dmmonhoc ON dmmonhoc.MaMonHoc = thuhocphi.MaMonHoc INNER JOIN defhocphi ON defhocphi.IDhocphi = thuhocphi.IDhocphi INNER JOIN dsgiaovien ON dsgiaovien.maGV = defhocphi.MaGV '

        if stat_ip == 1:
            cmd_sort += 'WHERE DATE(NgayThuHP) = "' + str(date_taken) + '" ;'
            print(cmd_sort)
        elif stat_ip == 2:
            cmd_sort += 'WHERE MONTH(NgayThuHP) = ' + str(month) + ' AND YEAR(NgayThuHP) = ' + str(year) + ' ;'
        else:
            cmd_sort += 'WHERE YEAR(NgayThuHP) = ' + str(year) + ' ;'

        print(cmd_sort)

        #Chạy CMD lấy dữ liệu
        run_cmd(cmd_sort)
        for row in recs:
            print(row)

        #Lưu trữ dữ liệu để xuất lên view
        opdata = []
        for row in recs:
            dt = datetime.datetime.strptime(str(row[2]),'%Y-%m-%d %H:%M:%S')
            #print(dt)

            per_rec = []
            for col in range(0,len(row)):
                per_rec.append(row[col])
            per_rec[2] = str(dt)
            #print(per_rec)

            opdata.append(per_rec)
     
        #print(opdata)
        
        #Load data onto view
        while self.hocphi_view.rowCount()>0:
            self.hocphi_view.removeRow(0)

        self.hocphi_view.setRowCount(len(opdata))
        self.hocphi_view.setColumnCount(7)
        
        self.header5 = self.hocphi_view.horizontalHeader()
        for col in range(0,7):
            self.header5.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        for row in range(0,len(opdata)):
            for col in range(0,7):
                self.hocphi_view.setItem(row,col,QTableWidgetItem(str(opdata[row][col])))


    #Def but ghi nhận thông tin đóng học phí
    def saveidhp(self):
        self.addids = self.idghinhan.text()
        if len(self.addids)>0:
            self.listids = []
            self.listids = self.addids.split(',')
            self.listids_int = []
            for item in self.listids:
                self.listids_int.append(int(item))

            #print(self.listids_int)

            allipdata = []
            for id in self.listids_int:
                cmd_ipidsinfo = 'SELECT IDhocphi, MaHS, MaMonHoc, HocPhi, MaGV FROM defhocphi WHERE IDhocphi = ' + str(id) + ';'
                print(cmd_ipidsinfo)
                run_cmd(cmd_ipidsinfo)
                per_recip = []
                for col in recs[0]:
                    per_recip.append(col)
                allipdata.append(per_recip)
            print(allipdata)

            #Tiến hành thủ tục nhập
            #Lấy max id thu học phí
            cmd_info = 'SELECT MAX(IDThuHP) FROM thuhocphi;'
            run_cmd(cmd_info)
            max_idthu = recs[0][0]
            #print(max_idthu)
            
            counter = max_idthu
            for id in range(0,len(self.listids_int)):
                #Input data vào database thu học phí
                counter += 1 
                print(id)
                cmd_ip = 'INSERT INTO thuhocphi VALUES (' + str(counter) + ', ' + str(allipdata[id][0]) + ', NOW() , "' + str(allipdata[id][1]) + '", "' + str(allipdata[id][2]) + '", "' + str(allipdata[id][1]) +'", ' + str(allipdata[id][0]) + ', "' + str(allipdata[id][2]) + '");'
                #print(cmd_ip)
                run_cmd(cmd_ip)

                #Cập nhật counter của defhocphi
                cmd_updatedefhp = 'UPDATE defhocphi SET SoLanDaDongTien = SoLanDaDongTien + 1 WHERE IDhocphi = ' + str(allipdata[id][0]) + ';'
                print(cmd_updatedefhp)
                run_cmd(cmd_updatedefhp)

    #Def button xóa id ghi nhận nhầm
    def xoaidnham(self):
        self.listidxoa = self.idxoa.text().split(',')
        if len(self.listidxoa)>0:
            for record in self.listidxoa:
                
                #Cập nhật lại số lần đóng tiền trong bảng defhocphi
                #Lấy IDhocphi
                cmd_getid = 'SELECT IDhocphi FROM thuhocphi WHERE IDThuHP = ' + record + ';'
                #print(cmd_getid)
                run_cmd(cmd_getid)
                idhp = recs[0][0]
                cmd_updatedefhp = 'UPDATE defhocphi SET SoLanDaDongTien = SoLanDaDongTien - 1 WHERE IDhocphi = ' + str(idhp) + ';'
                #print(cmd_updatedefhp)
                run_cmd(cmd_updatedefhp)



                #Xóa record chứa idthuhocphi trong bảng thuhocphi
                cmd_call = 'DELETE FROM thuhocphi WHERE IDThuHP = ' + record + ' ;' 
                print(cmd_call)
                run_cmd(cmd_call)










#This part intended left blank for easy recognize define a new window section





#Window 4 - Dialog Phân môn - nhóm - thời khóa biểu
class OpWinM4_pmntkb(Ui_phanmonnhomtkb):

    #Khởi tạo
    def setupUi(self,window_4):
        super().setupUi(window_4)

        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.lkthemxoanhom.clicked.connect(self.op_lkthemxoanhom)
        self.cnhstn.clicked.connect(self.op_themxoahs)
        self.dtkbn.clicked.connect(self.op_tdghlh)

        #Show current defined window
        window_4.show()

    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_4.hide()

    #Def nút mở window thêm xóa nhóm
    def op_lkthemxoanhom(self):
        self.ui = OpWinM41_lktxnhom()
        self.ui.setupUi(window_41)
        window_4.hide()


    #Def nút mở window thêm/bớt học sinh vào nhóm
    def op_themxoahs(self):
        self.ui = OpWinM42_tdtthstn()
        self.ui.setupUi(window_42)
        window_4.hide()

    #Def nút mở window thay đổi giờ học lịch học
    def op_tdghlh(self):
        self.ui = OpWinM43_tdghlh()
        self.ui.setupUi(window_43)
        window_4.hide()


#Window 4 - 1 - Dialog liệt kê , thêm và xóa nhóm
class OpWinM41_lktxnhom(Ui_lkthemxoanhom):

    #On Init
    def __init__(self):
        #Khởi tạo danh sách giáo viên
        cmd_dsgv = 'SELECT * FROM dsgiaovien;'
        run_cmd(cmd_dsgv)
        self.dsgv = ['0',]
        for row in recs:
            self.dsgv.append(row[0]+' - '+row[1])
            

        #Khởi tạo môn học
        cmd_dmmh = 'SELECT * FROM dmmonhoc;'
        run_cmd(cmd_dmmh)
        self.dmmh = ['0',]
        for row in recs:
            self.dmmh.append(row[0]+' - '+row[1])

    #Khởi tạo
    def setupUi(self,window_41):
        super().setupUi(window_41)

        #Init definition
        self.option_gv.addItems(self.dsgv)
        self.option_monhoc.addItems(self.dmmh)
        self.option_mamonthem.addItems(self.dmmh)
        self.option_gvphutrach.addItems(self.dsgv)


        #Button click definition
        self.hoantat.clicked.connect(self.ql)
        self.loc.clicked.connect(self.but_loc)
        self.taothemnhom.clicked.connect(self.ttn)
        self.nhapnhommoi.clicked.connect(self.nnm)
        self.xoanhom.clicked.connect(self.xoanhomcu)
     

        #Show current window
        window_41.show()



    #Def nút hoàn tất
    def ql(self):
        window_4.show()
        window_41.hide()


    #Def nút lọc
    def but_loc(self):
        ipdata = []
        
        #Nhập lớp
        ipdata.append(self.option_lop.currentText())
        
        #Nhập mã môn học
        ipmonhoc = self.option_monhoc.currentText()
        if len(ipmonhoc)>1:
            ipmonhoc = ipmonhoc[:4]
        ipdata.append(ipmonhoc)
        
        #Nhập mã giáo viên phụ trách nhóm
        ipmagv = self.option_gv.currentText()
        if len(ipmagv)>1:
            ipmagv = ipmagv[:4]
        ipdata.append(ipmagv)

        #Nhập ngày thứ cho thời khóa biểu
        ipthu = self.option_thu.currentText()
        if ipthu != 'Không chọn':
            ipthu_int = self.chg_date(ipthu)
            ipdata.append(ipthu_int)
        else:
            noneip = int(-1)
            ipdata.append(noneip)
        
        print(ipdata)

        #Tạo truy vấn lấy dữ liệu lên view
        cmd_sort = 'SELECT dsnhom.MaNhom, dsnhom.TenNhom, dsnhom.MaMonHoc, dsgiaovien.tenGV FROM dsnhom INNER JOIN dsgiaovien ON dsnhom.maGV = dsgiaovien.maGV INNER JOIN hstheonhom ON dsnhom.MaNhom = hstheonhom.MaNhom INNER JOIN thoikhoabieu ON dsnhom.MaNhom = thoikhoabieu.MaNhom WHERE thoikhoabieu.TinhTrangLich = 1'

        if ipdata[0] != '0':
            if len(ipdata[0]) == 1:
                itemloclop = '0' + ipdata[0]
            else:
                itemloclop = ipdata[0]
            cmd_sort += ' AND MID(dsnhom.MaNhom,5,2) = "' + itemloclop + '"'
        if len(ipdata[1]) > 1:
            cmd_sort += ' AND dsnhom.MaMonHoc = "' + ipdata[1] + '"'
        if len(ipdata[2]) > 1:
            cmd_sort += ' AND dsnhom.maGV = "' + ipdata[2] + '"'
        if ipdata[3] >= 0:
            cmd_sort += ' AND thoikhoabieu.Thu = ' + str(ipdata[3]) 
            
        cmd_sort += ';'

        print(cmd_sort)

        run_cmd(cmd_sort)
        #for row in recs:
        #    print(row)
        
        raw_export_data = []
        for row in recs:
            if row not in (raw_export_data):
                raw_export_data.append(row)

        for row in raw_export_data:
            print(row)

        #Lọc dữ liệu học sinh theo nhóm 
        all_group_stu = []
        for item in raw_export_data:
            cmd_sort =cmd_sort = 'SELECT dshocsinh.HoTenHS FROM dshocsinh INNER JOIN hstheonhom ON dshocsinh.MaHS = hstheonhom.MaHS WHERE MaNhom = "' + item[0] +'";'
            #print(cmd_sort)
            run_cmd(cmd_sort)

            each_group_stu = []
            for row in recs:
                #print(row[0])
                each_group_stu.append(row[0])
            all_group_stu.append(each_group_stu)
        
        print(all_group_stu)
        
        #Lấy dữ liệu toàn bộ học sinh cho view
        allstu_str = []
        for record in all_group_stu:
            eachstu_str = ''
            for item in record:
                eachstu_str += item + ', '
            eachstu_str = eachstu_str[:len(eachstu_str)-2] + '.'
            
            allstu_str.append(eachstu_str)

        print(allstu_str)

        #Lấy dữ liệu thời khóa biểu học cho view
        
        ip_alltkb = []
        for row in raw_export_data:
            cmd_sort = 'SELECT Thu FROM thoikhoabieu WHERE TinhTrangLich = 1 AND MaNhom = "' + row[0] + '";'
            #print(cmd_sort)
            run_cmd(cmd_sort)
            ip_tkbrow = ''
            for row in recs:
                ip_tkbrow += self.chg_wd(row[0]) + ' - '
            ip_tkbrow = ip_tkbrow[:len(ip_tkbrow)-3] + '.'
            #print(ip_tkbrow)
            ip_alltkb.append(ip_tkbrow)

        print(ip_alltkb)


        #Đẩy dữ liệu lọc lên table view
        self.dsnhomplus_view.setRowCount(len(raw_export_data))
        self.dsnhomplus_view.setColumnCount(6)

        self.header6 = self.dsnhomplus_view.horizontalHeader()
        for col in range(0,6):
            self.header6.setSectionResizeMode(col,QHeaderView.ResizeToContents)    

        for row in range(len(raw_export_data)):
            for col in range(0,4):
                self.dsnhomplus_view.setItem(row,col,QTableWidgetItem(str(raw_export_data[row][col])))
            self.dsnhomplus_view.setItem(row,4,QTableWidgetItem(allstu_str[row]))
            self.dsnhomplus_view.setItem(row,5,QTableWidgetItem(ip_alltkb[row]))



    # Def đổi thứ cho but_loc
    def chg_date(self,ipstring):
        if ipstring == 'Thứ hai':
            return 0
        if ipstring == 'Thứ ba':
            return 1
        if ipstring == 'Thứ tư':
            return 2
        if ipstring == 'Thứ năm':
            return 3
        if ipstring == 'Thứ sáu':
            return 4
        if ipstring == 'Thứ bảy':
            return 5
        if ipstring == 'Chủ nhật':
            return 6

    #Def đổi weekday thành thứ
    def chg_wd(self,ipint):
        if ipint == 0:
            return 'Thứ hai'
        if ipint == 1:
            return 'Thứ ba'
        if ipint == 2:
            return 'Thứ tư'
        if ipint == 3:
            return 'Thứ năm'
        if ipint == 4:
            return 'Thứ sáu'
        if ipint == 5:
            return 'Thứ bảy'
        if ipint == 6:
            return 'Chủ nhật'

    #Def nút tạo bảng thêm nhóm
    def ttn(self):
        #Lấy data input bao gồm số lượng nhóm cần tạo thêm và môn thêm
        ip_sln = self.option_slnhom.value()
        ip_monthem = self.option_mamonthem.currentText()
        ip_gvpt = self.option_gvphutrach.currentText()
        ip_lop = int(self.option_lopthem.currentText())
        if ip_lop < 10:
            malopthem = ip_monthem[:4] + '0' + str(ip_lop)
        else:
            malopthem = ip_monthem[:4] + str(ip_lop)
        malopthem = malopthem.replace(' ','')
        
        #Lấy năm hiện hành tạo mã lớp
        now = datetime.datetime.now()
        #print(type(now.year))
        malopthem += str(now.year)
        
        
        print(str(ip_sln) + ' - ' + ip_monthem + ' - ' + ip_gvpt + ' - ' + malopthem)
        cmd_maxip = 'SELECT MAX(SUBSTRING(MaNhom,length(MaNhom)-1,2)) FROM dsnhom WHERE SUBSTRING(MaNhom,1,length(MaNhom)-2) = "' + malopthem + '";'
        print(cmd_maxip)
        run_cmd(cmd_maxip)

        #print(recs)
        csmax = recs[0][0]
        print(csmax)
        print(type(csmax))
        if csmax is None:
            csmax = 0
        else:
            csmax = int(csmax)

        #Tạo bảng nhập dữ liệu
        self.dsnhomplus_view.setRowCount(ip_sln)
        self.dsnhomplus_view.setColumnCount(4)
        self.header6 = self.dsnhomplus_view.horizontalHeader()
        for col in range(0,4):
            self.header6.setSectionResizeMode(col,QHeaderView.ResizeToContents)
        
        self.infonewgroups = []
        
        for row in range(ip_sln):

            info_thisgroup = []
           

            self.dsnhomplus_view.setItem(row,2,QTableWidgetItem(ip_monthem))
            self.dsnhomplus_view.setItem(row,3,QTableWidgetItem(ip_gvpt))

            #Chèn mã nhóm vào view
            csmax += 1
            manhommoi = malopthem
            if csmax < 10:
                manhommoi += '0' + str(csmax)
            else:
                manhommoi += str(csmax)

            self.dsnhomplus_view.setItem(row,0,QTableWidgetItem(manhommoi))

            #Tạo và chèn tên nhóm
            mamonchen = manhommoi[:4]
            lopchen = manhommoi[4:6]
            namchen = manhommoi[6:10]
            sttchen = manhommoi[10:]
            print(mamonchen + ' - ' + lopchen + ' - ' + namchen + ' - ' + sttchen)
            tennhommoi = ''
            
            #Tạo tên môn vào tên nhóm mới
            if mamonchen == 'MATH':
                tennhommoi += 'Toán '
            elif mamonchen == 'ENGL':
                tennhommoi += 'Tiếng Anh '
            elif mamonchen == 'CHEM':
                tennhommoi += 'Hóa Học '
            elif mamonchen == 'INFT':
                tennhommoi += 'IT '

            
            #Tạo lớp vào tên nhóm
            tennhommoi += lopchen + ' '

            #Tạo niên khóa vào tên nhóm
            tennhommoi += 'niên khóa ' + namchen + ' '

            #Tạo thứ tự nhóm vào tên nhóm
            tennhommoi += 'Nhóm ' + sttchen

            print(tennhommoi)

            #Đưa tên nhóm lên bảng dữ liệu
            self.dsnhomplus_view.setItem(row,1,QTableWidgetItem(tennhommoi))

            #Chuẩn bị dữ liệu đầu vào để nhập
            info_thisgroup.append(manhommoi)
            info_thisgroup.append(tennhommoi)
            info_thisgroup.append(ip_monthem[:4])
            info_thisgroup.append(ip_gvpt[:4])

            self.infonewgroups.append(info_thisgroup)

    #Def nút thêm nhóm mới vào cơ sở dữ liệu
    def nnm(self):
        print(self.infonewgroups)
        for row in self.infonewgroups:
            cmd_inputdata = 'INSERT INTO dsnhom VALUES ("' + row[0] + '", "' + row[1] + '", "' + row[2] + '", "' + row[3] + '", "' + row[2] + '", "' + row[3] + '",1);'
            print(cmd_inputdata)
            run_cmd(cmd_inputdata)



    #Def nút xóa nhóm cũ khỏi cơ sở dữ liệu
    def xoanhomcu(self):
        info_nhomxoa = self.option_xoanhom.text()
        cacnhomxoa = info_nhomxoa.split(',')
        print(cacnhomxoa)
        for row in cacnhomxoa:
            cmd_xoa = 'DELETE FROM dsnhom WHERE MaNhom = "' + row + '";'
            print(cmd_xoa)
            run_cmd(cmd_xoa)


#Window 4 - 2 - Dialog thay đổi thông tin học sinh trong nhóm
class OpWinM42_tdtthstn(Ui_thaydoitthstrongnhom):

    #Khởi tạo
    def setupUi(self,window_42):
        super().setupUi(window_42)

        #Button clicked definition
        self.hoantat.clicked.connect(self.but_hoantat)
        self.locmanhommoi.clicked.connect(self.but_locnhommoi)
        self.dstoanbonhom.clicked.connect(self.but_fullnhom)
        self.viewhs.clicked.connect(self.hienthihs)
        self.ths.clicked.connect(self.themhsmoi)
        self.xoahs.clicked.connect(self.xoahscu)
        

        #Hiện window
        window_42.show()


    #Def nút hoàn tất
    def but_hoantat(self):
        window_4.show()
        window_42.hide()


    #Def nút hiển thị danh sách học sinh
    def hienthihs(self):
        manhomview = self.option_manhommoi.currentText()
        cmd_hienthi = 'SELECT dsnhom.MaNhom, hstheonhom.MaHS, dshocsinh.HoTenHS, dsnhom.maGV, dsgiaovien.tenGV FROM dsnhom INNER JOIN hstheonhom ON dsnhom.MaNhom = hstheonhom.MaNhom INNER JOIN dshocsinh ON hstheonhom.MaHS = dshocsinh.MaHS INNER JOIN dsgiaovien ON dsnhom.maGV = dsgiaovien.maGV WHERE dsnhom.MaNhom = "' + manhomview + '";'
        print(cmd_hienthi)
        run_cmd(cmd_hienthi)

        ipinfohs = []
        for row in recs:
            eachrow = []
            for item in row:
                eachrow.append(item)
            ipinfohs.append(eachrow)
        print(ipinfohs)
        
        #Upload thông tin lên bảng view

        #Reset bảng view trước mỗi lần hiển thị
        while (self.dshocsinhtheonhom_view.rowCount()>0):
            self.dshocsinhtheonhom_view.removeRow(0)

        #Tạo bảng view
        self.dshocsinhtheonhom_view.setRowCount(len(ipinfohs))
        self.dshocsinhtheonhom_view.setColumnCount(4)
        self.header7 = self.dshocsinhtheonhom_view.horizontalHeader()
        for col in range(0,4):
            self.header7.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Thêm dữ liệu vào bảng view
        for row in range(0,len(ipinfohs)):
            self.dshocsinhtheonhom_view.setItem(row,0,QTableWidgetItem(ipinfohs[row][1]))
            self.dshocsinhtheonhom_view.setItem(row,1,QTableWidgetItem(ipinfohs[row][2]))
            self.dshocsinhtheonhom_view.setItem(row,2,QTableWidgetItem(ipinfohs[row][0]))
            self.dshocsinhtheonhom_view.setItem(row,3,QTableWidgetItem(ipinfohs[row][4]))

        #Update danh sách học sinh thuộc và không thuộc nhóm
        stu_in_group =  []
        instu_code = []
        for item in range(0,len(ipinfohs)):
            each_stu = []
            each_stu.append(ipinfohs[item][1])
            each_stu.append(ipinfohs[item][2])
            stu_in_group.append(each_stu)
            instu_code.append(ipinfohs[item][1])

        print(stu_in_group)
        print(instu_code)

        cmdsort_allstu = 'SELECT MaHS FROM dshocsinh;'
        run_cmd(cmdsort_allstu)
        #print(recs)
        allstu_code = []
        for row in recs:
            allstu_code.append(row[0])
        #print(allstu_code)

        outstu_code = [item for item in allstu_code if item not in instu_code]
        print(outstu_code)

        #Load danh sách học sinh lên combobox (instu_code & outstu_code)
        self.option_chonhsthem.clear()
        self.option_chonhsthem.addItems(outstu_code)
        
        self.option_chonhsxoa.clear()
        self.option_chonhsxoa.addItems(instu_code)

    #Def nút thêm học sinh mới vào nhóm 
    def themhsmoi(self):
        mahsthem = self.option_chonhsthem.currentText()
        print(mahsthem)
        manhomthem = self.option_manhommoi.currentText()
        cmd_getmaxid = 'SELECT MAX(IDHSTheoNhom) FROM hstheonhom;'
        run_cmd(cmd_getmaxid)
        max_id = recs[0][0]
        print(max_id)
        max_id += 1
        cmd_them = 'INSERT INTO hstheonhom VALUES(' + str(max_id) + ', "' + mahsthem + '", "' + manhomthem + '", "' + mahsthem + '");'
        print(cmd_them)
        run_cmd(cmd_them)





    #Def nút xóa học sinh cũ khỏi nhóm
    def xoahscu(self):
        mahsxoa = self.option_chonhsxoa.currentText()
        print(mahsxoa)
        cmd_xoa = 'DELETE FROM hstheonhom WHERE MaHS = "' + mahsxoa + '";'
        print(cmd_xoa)
        run_cmd(cmd_xoa)


    #Def nút hiển thị danh sách toàn bộ nhóm
    def but_fullnhom(self):
        cmd_mn_indsnhom = 'SELECT MaNhom FROM dsnhom;'
        run_cmd(cmd_mn_indsnhom)
        mn_indsnhom = []
        for row in recs:
            mn_indsnhom.append(row[0])

        self.option_manhommoi.clear()
        self.option_manhommoi.addItems(mn_indsnhom)


    #Def nút lấy nhóm chưa có học sinh
    def but_locnhommoi(self):

        cmd_mn_indsnhom = 'SELECT MaNhom FROM dsnhom;'
        run_cmd(cmd_mn_indsnhom)
        mn_indsnhom = []
        for row in recs:
            mn_indsnhom.append(row[0])
        
        cmd_mn_inhstheonhom = 'SELECT MaNhom FROM hstheonhom;'
        run_cmd(cmd_mn_inhstheonhom)
        mn_inhstheonhom = []
        for row in recs:
            if not (row[0] in mn_inhstheonhom):
                mn_inhstheonhom.append(row[0])
                
        print(mn_indsnhom)
        print(mn_inhstheonhom)
        op_manhom = list(set(mn_indsnhom)-set(mn_inhstheonhom))
        print(op_manhom)
        self.option_manhommoi.clear()
        self.option_manhommoi.addItems(op_manhom)


#Window 4 - 3 - Dialog thay đổi giờ học lịch học
class OpWinM43_tdghlh(Ui_thaydoighlh):
    def __init__(self):
        cmd_chonnhom = 'SELECT MaNhom FROM dsnhom;'
        run_cmd(cmd_chonnhom)
        self.allgroup = []
        for row in recs:
            self.allgroup.append(row[0])
        print(self.allgroup)

        group_intkb = []
        cmd_nhomintkb = 'SELECT MaNhom FROM thoikhoabieu WHERE TinhTrangLich = 1;'
        run_cmd(cmd_nhomintkb)
        for row in recs:
            if row[0] not in group_intkb:
                group_intkb.append(row[0])

        self.group_notintkb = [item for item in self.allgroup if item not in group_intkb]

    #Khởi tạo
    def setupUi(self,window_43):
        super().setupUi(window_43)

        #Khởi tạo combobox
        self.nhom.addItems(self.allgroup)

        #Button clicked definition
        self.hoantat.clicked.connect(self.but_hoantat)
        self.locall.clicked.connect(self.but_allgroup)
        self.locchuatkb.clicked.connect(self.but_groupnotintkb)
        self.themgiohoc.clicked.connect(self.addtime)
        self.capnhat.clicked.connect(self.onupdate_nhom)
        self.xoagiohoc.clicked.connect(self.removetkb)
        self.capnhatgiohoc.clicked.connect(self.updatetkb)

        #Hiện window
        window_43.show()

        #Lệnh chạy cập nhật học sinh trong nhóm
        self.nhom.currentIndexChanged.connect(self.onupdate_nhom)

    #Def nút chọn toàn bộ nhóm
    def but_allgroup(self):
        self.nhom.clear()
        self.nhom.addItems(self.allgroup)

    #Def nút chọn nhóm chưa có thời khóa biểu
    def but_groupnotintkb(self):
        self.nhom.clear()
        self.nhom.addItems(self.group_notintkb)


    #Def nút hoàn tất
    def but_hoantat(self):
        window_4.show()
        window_43.hide()

    #Def khi cập nhật lựa chọn nhóm
    def onupdate_nhom(self):

        #Def dòng hiển thị học sinh trong nhóm
        print(self.nhom.currentText())
        cmd_opstus = 'SELECT hstheonhom.MaHS, dshocsinh.HoTenHS FROM hstheonhom INNER JOIN dshocsinh ON hstheonhom.MaHS = dshocsinh.MaHS WHERE hstheonhom.MaNhom = "' + self.nhom.currentText() + '";'
        print(cmd_opstus)
        run_cmd(cmd_opstus)
        opstu = []
        for row in recs:
            opstu.append(row[1])
        
        #print(opstu)

        opstu_forqline = ''
        for item in opstu:
            opstu_forqline += item + ', '
        opstu_forqline = opstu_forqline[:len(opstu_forqline)-2] + '.'
        print(opstu_forqline)
        self.hstrongnhom_view.clear()
        self.hstrongnhom_view.setText(opstu_forqline)

        #Cập nhật list id giờ học lên các combobox
        cmd_ipid = 'SELECT IDTKB, NgayGioHoc, Thu FROM thoikhoabieu WHERE TinhTrangLich = 1 AND MaNhom = "' + self.nhom.currentText() + '";'
        run_cmd(cmd_ipid)
        idlist = []
        alldata = []
        for row in recs:
            rowdata = []
            idlist.append(str(row[0]))
            rowdata.append(str(row[0]))
            rowdata.append(str(row[1]))
            rowdata.append(str(row[2]))
            alldata.append(rowdata)
        
        print(idlist)
        print(alldata)

        self.option_idxoa.clear()
        self.option_idxoa.addItems(idlist)

        self.option_idcapnhat.clear()
        self.option_idcapnhat.addItems(idlist)
        

        #Chuẩn bị bảng xuất dữ liệu
        self.thaydoigiohoc_view.setRowCount(len(alldata))
        self.thaydoigiohoc_view.setColumnCount(3)
        self.header8 = self.thaydoigiohoc_view.horizontalHeader()
        for col in range(0,3):
            self.header8.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Xuất dữ liệu ra bảng
        for row in range(0,len(alldata)):
            self.thaydoigiohoc_view.setItem(row,0,QTableWidgetItem(alldata[row][0]))
            self.thaydoigiohoc_view.setItem(row,1,QTableWidgetItem(self.chg_wd(int(alldata[row][2]))))
            dt = datetime.datetime.strptime(alldata[row][1],'%Y-%m-%d %H:%M:%S')
            time = str(dt.time())
            print(time)
            self.thaydoigiohoc_view.setItem(row,2,QTableWidgetItem(time))



    #Def chọn ngày gần với weekday nhất
    def opdatetime(self,opwd,optime):
        dt = datetime.datetime.now()
      
        opwd_int = self.chg_date(opwd)
        #print(opwd_int)
        #print(dt.weekday())
        
        while ((dt.weekday() != opwd_int) or (dt.hour != optime.hour) or (dt.minute != optime.minute)):
            #Chỉnh ngày
            while (dt.weekday() != opwd_int):
                ord_dt = dt.toordinal()
                ord_dt += 1
                dt = dt.fromordinal(ord_dt)
                #print(dt)
                #print(dt.weekday())


            #Chỉnh giờ
            hour_period = datetime.timedelta(hours = 1)
            #print(optime.hour)
            while dt.hour != optime.hour:
                dt += hour_period
                #print(dt)

            #Chỉnh phút
            min_period = datetime.timedelta(minutes = 1)
            #print(optime.minute)
            while dt.minute != optime.minute:
                dt += min_period
                #print(dt)


        return dt


    #Def thêm giờ học mới
    def addtime(self):
        buoihoc = self.option_1_chonbuoihoc.currentText()
        giohoc = self.option_1_giohoc.time().toPyTime()
        #print(buoihoc)
        #print(giohoc)
        need_dt = self.opdatetime(buoihoc,giohoc)
        print(need_dt)

        #Push record lên cơ sở dữ liệu
        cmd_maxid = 'SELECT MAX(IDTKB) FROM thoikhoabieu;'
        run_cmd(cmd_maxid)
        max_id = recs[0][0]
        max_id += 1
        print(str(max_id) + ' - ' + self.nhom.currentText() + ' - ' + str(need_dt))
        cmd_push = 'INSERT INTO thoikhoabieu VALUES(' + str(max_id) + ',"' + self.nhom.currentText() + '", "' + str(need_dt) + '", 1, ' + str(self.chg_date(buoihoc)) + ', "'+ self.nhom.currentText()  + '");'
        print(cmd_push)
        run_cmd(cmd_push)
        
        #Cập nhật lại view
        self.onupdate_nhom()

    #Def xóa giờ học cũ
    def removetkb(self):
        cmd_remove = 'DELETE FROM thoikhoabieu WHERE IDTKB = ' + self.option_idxoa.currentText() + ';'
        print(cmd_remove)
        run_cmd(cmd_remove)

        #Cập nhật lại view
        self.onupdate_nhom()

    #Def cập nhật giờ học đã có
    def updatetkb(self):
        idip = self.option_idcapnhat.currentText()
        buoihoc = self.option_2_buoihoc.currentText()
        giohoc = self.option_2_giohoc.time().toPyTime()
        need_dt = self.opdatetime(buoihoc,giohoc)
        print(need_dt)
        cmd_update = 'UPDATE thoikhoabieu SET NgayGioHoc = "' + str(need_dt) + '", Thu = ' + str(self.chg_date(buoihoc)) + ' WHERE MaNhom = "' + self.nhom.currentText() + '" AND IDTKB = ' + str(idip) + ' ;'
        print(cmd_update)
        run_cmd(cmd_update)

        #Cập nhật lại view
        self.onupdate_nhom()



    # Def đổi thứ cho but_loc
    def chg_date(self,ipstring):
        if ipstring == 'Thứ hai':
            return 0
        if ipstring == 'Thứ ba':
            return 1
        if ipstring == 'Thứ tư':
            return 2
        if ipstring == 'Thứ năm':
            return 3
        if ipstring == 'Thứ sáu':
            return 4
        if ipstring == 'Thứ bảy':
            return 5
        if ipstring == 'Chủ nhật':
            return 6

    #Def đổi weekday thành thứ
    def chg_wd(self,ipint):
        if ipint == 0:
            return 'Thứ hai'
        if ipint == 1:
            return 'Thứ ba'
        if ipint == 2:
            return 'Thứ tư'
        if ipint == 3:
            return 'Thứ năm'
        if ipint == 4:
            return 'Thứ sáu'
        if ipint == 5:
            return 'Thứ bảy'
        if ipint == 6:
            return 'Chủ nhật'




#This part intended left blank



#Window 5 - Dialog Phân tích lợi nhuận
class OpWinM5_ptln(Ui_phantichloinhuan):
    
    #Khởi tạo
    def setupUi(self,window_5):
        super().setupUi(window_5)

        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.bclgv.clicked.connect(self.opbclgv)
        self.capnhatthongtin.clicked.connect(self.cntt)
        self.chiluonggv.clicked.connect(self.ctlgv)

        #Deact các option chưa code
        self.lr.setEnabled(False)
        self.bctq.setEnabled(False)
        self.bccptt.setEnabled(False)

        #Deact các options chỉ dùng cho bản quản trị chứ không phải bản nhập liệu
        self.chiluonggv.setEnabled(False)
        self.capnhatthongtin.setEnabled(False)


        #Show current window
        window_5.show() 


    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_5.hide()

    #Def nút mở bảng chi lương giáo viên
    def opbclgv(self):
        self.ui = OpWinM51_bclgv()
        self.ui.setupUi(window_51)
        window_5.hide()

    #Def nút mở bảng cập nhật thông tin
    def cntt(self):
        self.ui = OpWinM52_cnttluong()
        self.ui.setupUi(window_52)
        window_5.hide()

    #Def nút mở bảng chi trả lương giáo viên
    def ctlgv(self):
        self.ui = OpWinM53_lctluonggv()
        self.ui.setupUi(window_53)
        window_5.hide()



class OpWinM51_bclgv(Ui_bclgiaovien):

    #Def init
    def __init__(self):
        dtnow = datetime.date.today()
        print(dtnow)

        self.dtcurmonth = dtnow.month
        self.dtcuryear = dtnow.year
        print(self.dtcurmonth,' - ',self.dtcuryear)

        #Lấy items gửi vào lựa chọn
        cmd_mgv = 'SELECT maGV,tenGV FROM dsgiaovien;'
        run_cmd(cmd_mgv)
        self.allmgv = []
        for row in recs:
            self.allmgv.append(row[0] + '-' + row[1])



    #Khởi tạo
    def setupUi(self,window_51):
        super().setupUi(window_51)

        #Button clicked definition
        self.hoantat.clicked.connect(self.ql)
        self.loc.clicked.connect(self.whileloc)

        #Define mặc định cho spinbox tháng năm
        self.thang.setValue(self.dtcurmonth)
        self.nam.setValue(self.dtcuryear)

        #Xử lý khi lựa chọn option
        self.locoption.toggled.connect(self.whilelocoption)
        self.locall.toggled.connect(self.whilelocall)

        #Chèn danh sách mã giáo viên vào list trước
        self.magv.addItems(self.allmgv)

        #Define trước khi option được chọn
        self.magv.setEnabled(False)

        #Show current window
        window_51.show()


    #Def nút lọc all
    def whilelocall(self):
        self.magv.setEnabled(False)

    #Def nút lọc riêng mã giáo viên
    def whilelocoption(self):
        self.magv.setEnabled(True)

    #Def nút quay lại
    def ql(self):
        window_5.show()
        window_51.hide()

    #Def nút chuyển số integer thành chuỗi có dot seperate
    def viewnum(self,ipnum):
        strnum = str(ipnum)
        #print(strnum)
        liststr = []
        while len(strnum)>0:
            if len(strnum) >= 3:
                #print(strnum[-4:-1])
                liststr.append(strnum[len(strnum)-3:])
                strnum = strnum[:len(strnum)-3]
                print(strnum)
                
            else:
                liststr.append(strnum)
                strnum = ''

        opstr = ''
        for item in liststr:
          opstr = item + '.' + opstr
        
        opstr = opstr[:-1]

        return opstr

    #Def nút lọc
    def whileloc(self):
        mgv = []
        if self.locoption.isChecked() == True:
            ip_onlymgv = self.magv.currentText()
            ipdata = ip_onlymgv.split('-')
            mgv.append(ipdata[0])
            #print(mgv)
        else:
            for item in self.allmgv:
                ip_mgv = []
                ip_mgv = item.split('-')
                mgv.append(ip_mgv[0])

        #print(mgv)

        #Bắt đầu lọc
        cmd_ip = 'SELECT * FROM dsgiaovien;'
        threading.Thread(run_cmd(cmd_ip))
        #print(recs)
        data = []
        for item in recs:
            if item[0] in mgv:
                data.append(item)

        #print(data)

        #Build thêm dữ liệu cập nhật trạng thái lương hiện tại để tất toán thu chi







        #Đẩy dữ liệu lên data
        self.payment_view.setRowCount(len(data))
        self.payment_view.setColumnCount(8)

        self.header16 = self.payment_view.horizontalHeader()
        for col in range(0,8):
            self.header16.setSectionResizeMode(col,QHeaderView.ResizeToContents)
        
        print(data)

        for row in range(0,len(data)):

            #Xử lý khởi tạo trước biến amoutwithrate để tránh trường hợp biến không được khai báo khi dòng if không được gọi
            amountwithrate = 0
            
            #Lấy dữ liệu từ bảng thuhocphi với điều kiện mã giáo viên & ngày tháng tương ứng
            cmd_takelistinfo = 'SELECT IDThuHP FROM thuhocphi WHERE MONTH(NgayThuHP) = ' + str(self.thang.value()) + ' AND YEAR(NgayThuHP)=' + str(self.nam.value()) + ';'
            #print(cmd_takelistinfo)
            threading.Thread(run_cmd(cmd_takelistinfo))
            inmonth_paidrec = []
            for item in recs:
                inmonth_paidrec.append(item[0])
            #print(inmonth_paidrec)
            

            #Lấy dữ liệu tổng lương tính kèm rating
            cmd_sumtotalpaidwithrate = 'SELECT SUM(defhocphi.HocPhi) FROM defhocphi INNER JOIN thuhocphi ON thuhocphi.IDhocphi = defhocphi.IDhocphi WHERE defhocphi.MaGV = "'+ data[row][0] + '" AND MONTH(NgayThuHP) = ' + str(self.thang.value()) + ' AND YEAR(NgayThuHP)=' + str(self.nam.value()) + ';'
            #print(cmd_sumtotalpaidwithrate)
            threading.Thread(run_cmd(cmd_sumtotalpaidwithrate))
            incomeinmonth = recs[0][0]
            print(incomeinmonth)
            
            #Tính payment theo rate đã lưu in month



            #Cập nhật kèm dữ liệu tổng lương theo hệ số này vào bảng lương giáo viên trong tháng
            #print(isinstance(incomeinmonth,decimal.Decimal))

            #print(type(incomeinmonth))

            if isinstance(incomeinmonth,decimal.Decimal):
                cmd_updatepayment = 'UPDATE dsgiaovien SET TotalProfitAsRated = ' + str(incomeinmonth) + ' WHERE maGV = "' + data[row][0] + '";'
                print(cmd_updatepayment)
                threading.Thread(run_cmd(cmd_updatepayment))
                


            #Tính toán lương payment theo rating
            unzip1ststeps = []
            unzip1ststeps = data[row][2].split(',')
            print(unzip1ststeps)

            if len(unzip1ststeps[0])>2:


                unzip2ndsteps = []
                for item in unzip1ststeps:
                    sppart = []
                    sppart = item.split('-')
                    numpart = []
                    for it in range(0,2):
                        numpart.append(int(float(sppart[it])))
                    numpart.append(float(sppart[2]))
                    unzip2ndsteps.append(numpart)
               
                print(unzip2ndsteps)

                paysplit = []
                calcamount = incomeinmonth
                
                print(str(calcamount) + ' - ' + str(type(calcamount)))
                print(unzip2ndsteps)

                if not isinstance(calcamount,decimal.Decimal):
                    calcamount = 0

                for step in range(0,len(unzip2ndsteps)):
                    if (calcamount/1000000) >= unzip2ndsteps[step][1]:
                        paysplit.append(unzip2ndsteps[step][1])
                        calcamount -= unzip2ndsteps[step][1]*1000000
                    else:
                        paysplit.append(calcamount/1000000)
                        calcamount = 0

                print(paysplit)
                    
                #Tính total Payment
                amountwithrate = 0
                for rate in range(0,len(unzip2ndsteps)):
                    amountwithrate += float(paysplit[rate])*1000000*float(unzip2ndsteps[rate][2])

                
            print(amountwithrate)

            #Load lên bảng view
            self.payment_view.setItem(row,3,QTableWidgetItem(self.viewnum(int(amountwithrate))))
            print("Đã load lương theo hệ số cho mã giáo viên " + data[row][0] + "!!!!")

            #Push dữ liệu lên bảng view
            self.payment_view.setItem(row,0,QTableWidgetItem(data[row][0]))
            self.payment_view.setItem(row,1,QTableWidgetItem(data[row][1]))
            self.payment_view.setItem(row,2,QTableWidgetItem(data[row][2]))
            self.payment_view.setItem(row,4,QTableWidgetItem(self.viewnum(data[row][3])))

            
            #Tính tổng thu nhập
            totalpaid = int(amountwithrate) + data[row][3]
            self.payment_view.setItem(row,5,QTableWidgetItem(self.viewnum(totalpaid)))


            #Tổng lương đã ứng trước
            cmd_ungluong = 'SELECT SUM(PaidAmount) FROM traluonggv WHERE maGV = "' + data[row][0] + '" AND MONTH(PaidDate)=' + str(self.thang.value()) + ' AND YEAR(PaidDate)=' + str(self.nam.value()) + ';'
            #print(cmd_ungluong)
            threading.Thread(run_cmd(cmd_ungluong))
            op_paid = recs[0][0]
            #print(recs)
            
            #print(op_paid)

            #Load lên bảng dữ liệu , trước khi load kiểm tra tồn tại biến
            if op_paid is None:
                op_paid = 0
            elif op_paid is decimal.Decimal:
                op_paid = int(recs[0][0])
            print(str(op_paid) + ' - ' + str(type(op_paid)))
            
            self.payment_view.setItem(row,6,QTableWidgetItem(self.viewnum(op_paid)))

            luongcuoithang = totalpaid - op_paid

            self.payment_view.setItem(row,7,QTableWidgetItem(self.viewnum(luongcuoithang)))


#Bảng cập nhật thông tin lương giáo viên
class OpWinM52_cnttluong(Ui_capnhatttluong):

    #def init
    def __init__(self):

        #Lấy danh sách giáo viên để cập nhật
        cmd_ip = 'SELECT maGV, tenGV FROM dsgiaovien;'
        run_cmd(cmd_ip)
        self.listgv = []
        for item in recs:
            ipstr = item[0] + '-' + item[1]
            self.listgv.append(ipstr)
        print(self.listgv)

        

    #Khởi tạo
    def setupUi(self,window_52):
        super().setupUi(window_52)

        #Chèn vào list chọn
        self.mgvblhs.addItems(self.listgv)
        self.mgvhsl.addItems(self.listgv)
        self.mgvlcd.addItems(self.listgv)
        self.mgvcnl.addItems(self.listgv)

        #Button clicked definition
        self.hoantat.clicked.connect(self.ql)
        self.httt.clicked.connect(self.htthongtin)
        self.updatebl.clicked.connect(self.cnbacluong)
        self.capnhathsl.clicked.connect(self.thembacluong)
        self.delperhsl.clicked.connect(self.xoabacluong)


        #Khi click tự động cập nhật lại thông tin
        self.updatebl.clicked.connect(self.htthongtin)
        self.capnhathsl.clicked.connect(self.htthongtin)
        
        self.delperhsl.clicked.connect(self.htthongtin)
        self.delperhsl.clicked.connect(self.onupdate_magvxoabl)

        self.delallhsl.clicked.connect(self.xoaallhsl)
        self.delallhsl.clicked.connect(self.htthongtin)
        
        self.udnewbasepm.clicked.connect(self.new_basicpayment)
        self.udnewbasepm.clicked.connect(self.htthongtin)

        #Cập nhật khi combobox khi được chọn
        self.mgvcnl.currentIndexChanged.connect(self.mgvcnl_onupdate)
        self.mgvhsl.currentIndexChanged.connect(self.onupdate_magvxoabl)

        #Show current window
        window_52.show()

    #Def nút quay lại
    def ql(self):
        window_5.show()
        window_52.hide()

    #Def khi cập nhật mgv để update bậc lương
    def mgvcnl_onupdate(self):
        mgv_str = self.mgvcnl.currentText()
        mgv_list = mgv_str.split('-')
        mgv = mgv_list[0]
        #print(mgv)

        cmd_slt = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgv + '";'
        #print(cmd_slt)
        run_cmd(cmd_slt)
        oup_allsteps = recs[0][0]
        print(oup_allsteps)
        splitsteps = oup_allsteps.split(',')
        print(splitsteps)
        
        self.conn_steps = []
        for item in splitsteps:
            perit = item.split('-')
            perit[0] = str(int(perit[0]))
            for it in range(1,len(perit)): 
                perit[it] = float(perit[it])
            self.conn_steps.append(perit)
        print(self.conn_steps)

        self.steps_chose = []
        for item in self.conn_steps:
            self.steps_chose.append(item[0])
        print(self.steps_chose)

        self.bacsua.clear()
        self.bacsua.addItems(self.steps_chose)

    #Def chức năng split hệ số lương
    def split_hsl(self,mgvip):
        cmdtakehsl = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgvip + '";'
        threading.Thread(run_cmd(cmdtakehsl))
        print(recs)
        opraw_db = recs[0][0]
        opraw_str = opraw_db.split(',')
        cooked_data = []
        for item in opraw_str:
            per_raw = item.split('-')
            cooked_data.append(per_raw)

        #print(cooked_data)
        return cooked_data

    #Def chức năng split thứ tự hệ số lương
    def split_stt_hsl(self,mgvip):
        cmdtakehsl = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgvip + '";'
        threading.Thread(run_cmd(cmdtakehsl))
        print(recs)
        opraw_db = recs[0][0]
        opraw_str = opraw_db.split(',')
        cooked_data = []
        for item in opraw_str:
            per_raw = item.split('-')
            cooked_data.append(per_raw)
        print(cooked_data)

    #Def connect danh sách bậc lương
    def join_listpayment(self,iplist):
        split_inside = '-'
        split_perstep = ','

        side_list = []
        for item in iplist:
            perstep = split_inside.join(item)
            side_list.append(perstep)
        op_str = split_perstep.join(side_list)

        #print(op_str)
        return op_str

    #Def nút cập nhật bậc lương
    def cnbacluong(self):
        rawstr_mgvupdate = self.mgvcnl.currentText()
        mgvsplit = rawstr_mgvupdate.split('-')
        mgvupdate = mgvsplit[0]
        blupdate = self.bacsua.currentText()
        luongamount = self.luongbac.value()
        hsls = self.hslsua.value()
        print(mgvupdate,' - ',blupdate,' - ',luongamount,' - ',hsls)
        ds_steps_hsl = self.split_hsl(mgvupdate)
        print(ds_steps_hsl)
        ds_steps_oup = []
        for item in range(0,len(ds_steps_hsl)):
            peritem = []
            peritem.append(ds_steps_hsl[item][0])

            if ds_steps_hsl[item][0] == blupdate:
                amt_lg = str(luongamount)
                rate = str(hsls)
                peritem.append(amt_lg[:4])
                peritem.append(rate[:4])
                #print(peritem)
            else:
                peritem.append(ds_steps_hsl[item][1])
                peritem.append(ds_steps_hsl[item][2])


            ds_steps_oup.append(peritem)

        print(ds_steps_oup)
        print(self.join_listpayment(ds_steps_oup))

        cmd_updatestepspaid = 'UPDATE dsgiaovien SET PayRate = "' + self.join_listpayment(ds_steps_oup) + '" WHERE maGV = "' + mgvupdate + '";'
        print(cmd_updatestepspaid)
        run_cmd(cmd_updatestepspaid)

    #Def nút thêm bậc lương mới
    def thembacluong(self):
        mgv_ca = self.mgvblhs.currentText()
        mgv_split = mgv_ca.split('-')
        mgv_ca_cooked = mgv_split[0]

        step_ca = self.steppayment.value()
        step_ca_flop = float("{:.2f}".format(step_ca))
        hsl_ca = self.hsl.value()
        hsl_ca_flop = float("{:.2f}".format(hsl_ca))
        
        print(mgv_ca_cooked,' - ',step_ca_flop,' - ',hsl_ca_flop)

        cmd_loc = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgv_ca_cooked + '";'
        #print(cmd_loc)
        threading.Thread(run_cmd(cmd_loc))
        print(recs[0][0])
        check_rate = recs[0][0]
        if check_rate == '0':
            addid = 1
        else:
            raw_idlist = self.split_hsl(mgv_ca_cooked)
            print(raw_idlist)
            ip_rawlist = []
            for item in raw_idlist:
                ip_rawlist.append(int(item[0]))
            addid = max(ip_rawlist)+1

        #print(addid)
        oupstr = str(addid)+ '-' + str(step_ca_flop) + '-' + str(hsl_ca_flop)
        #print(oupstr)

        if check_rate == '0':
            check_rate = oupstr
        else:
            check_rate = check_rate + ',' + oupstr

        #print(check_rate)
        cmd_updatecr = 'UPDATE dsgiaovien SET PayRate = "' + check_rate + '" WHERE maGV = "' + mgv_ca_cooked + '";'
        #print(cmd_updatecr)
        threading.Thread(run_cmd(cmd_updatecr))

    #Def khi cập nhật mã giáo viên xóa bậc lương
    def onupdate_magvxoabl(self):
        op_hsl = self.mgvhsl.currentText()
        #print(op_hsl)
        list_mid = op_hsl.split('-')
        mgv_spt = list_mid[0]
        #print(mgv_spt)
        cmd_sltlist = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgv_spt + '";'
        threading.Thread(run_cmd(cmd_sltlist))
        print(recs[0][0])
        rawlst = recs[0][0]
        raw_1st = rawlst.split(',')
        #print(raw_1st)
        self.bacxoa.clear()
        self.bacxoa.addItems(raw_1st)

        if len(raw_1st) <= 1:
            self.delperhsl.setEnabled(False)
        else:
            self.delperhsl.setEnabled(True)


    #Def nút xóa bậc lương
    def xoabacluong(self):
        infoxoa = self.bacxoa.currentText()
        mgv_raw = self.mgvhsl.currentText()
        mgv_list = mgv_raw.split('-')
        mgv = mgv_list[0]
        print(mgv,' - ',infoxoa)

        cmd_sltstr = 'SELECT PayRate FROM dsgiaovien WHERE maGV = "' + mgv + '";'
        #print(cmd_sltstr)
        threading.Thread(run_cmd(cmd_sltstr))
        #print(recs[0][0])
        ip_stepsdata = recs[0][0]
        #print(defitem)
        raw_stepsdata = ip_stepsdata.split(',')
        #print(raw_stepsdata)
        raw_stepsdata.remove(infoxoa)
        #print(raw_stepsdata)
        cooked_stepsdata = ','.join(raw_stepsdata)
        print(cooked_stepsdata)

        #Update PayRate in dsgiaovien
        cmd_update = 'UPDATE dsgiaovien SET PayRate = "' + cooked_stepsdata + '" WHERE maGV = "' + mgv + '";'
        print(cmd_update)
        threading.Thread(run_cmd(cmd_update))

    #Def nút xóa toàn bộ hệ số lương
    def xoaallhsl(self):
        raw_mgv = self.mgvhsl.currentText()
        mgv_list = raw_mgv.split('-')
        mgv = mgv_list[0]
        #print(mgv)
        cmd_del = 'UPDATE dsgiaovien SET PayRate = "0" WHERE maGV = "' + mgv + '";'
        print(cmd_del)
        threading.Thread(run_cmd(cmd_del))

    #Def nút cập nhật mức lương cố định mới
    def new_basicpayment(self):
        mgv_raw = self.mgvlcd.currentText()
        mgv_list = mgv_raw.split('-')
        mgv = mgv_list[0]
        new_basepayment = int(round(self.basepayment.value(),2)*1000000)
        print(mgv,'-',new_basepayment)
        cmd_updatelc = 'UPDATE dsgiaovien SET BasePayment = ' + str(new_basepayment) + ' WHERE maGV = "' + mgv + '";'
        print(cmd_updatelc)
        threading.Thread(run_cmd(cmd_updatelc))

    #Def chuyển số integer thành chuỗi có dot seperate
    def viewnum(self,ipnum):
        strnum = str(ipnum)
        #print(strnum)
        liststr = []
        while len(strnum)>0:
            if len(strnum) >= 3:
                #print(strnum[-4:-1])
                liststr.append(strnum[len(strnum)-3:])
                strnum = strnum[:len(strnum)-3]
                #print(strnum)
                
            else:
                liststr.append(strnum)
                strnum = ''

        opstr = ''
        for item in liststr:
          opstr = item + '.' + opstr
        
        opstr = opstr[:-1]

        return opstr

    #Def nút hiển thị thông tin
    def htthongtin(self):
        #Lấy thông tin để cập nhật vào bảng view
        cmd_srt = 'SELECT maGV, tenGV, PayRate, BasePayment, TotalProfitAsRated, PaymentAfterRate FROM dsgiaovien;'
        print(cmd_srt)
        threading.Thread(run_cmd(cmd_srt))
        print(recs)

        #Cập nhật thông tin lên bảng view
        while (self.payment_view.rowCount()>0):
            self.payment_view.removeRow(0)

        self.payment_view.setRowCount(len(recs))
        self.payment_view.setColumnCount(8)
        
        self.header17 = self.payment_view.horizontalHeader()
        for col in range(0,8):
            self.header17.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Chép recs ra dữ liệu thường
        view_list = []

        for row in range(0,len(recs)):
            peritem = []
            for col in range(0,6):
                peritem.append(recs[row][col])
            view_list.append(peritem)
        
        #Tính tổng thu đã tạo cho trung tâm
        for row in range(0,len(view_list)):
            view_list[row][5] = view_list[row][3] + view_list[row][4]


        #Tính lương được rated
        rate_list = []
        for row in range(0,len(view_list)):
            rate_list.append(view_list[row][2])
        #print(rate_list)

        #Cắt chuỗi rate lần 1
        sep_turn1st = []
        for item in rate_list:
            per_it = item.split(',')
            sep_turn1st.append(per_it)
        print(sep_turn1st)        

        #Cắt chuỗi rate lần 2 
        sep_turn2nd = []
        for item in sep_turn1st:
            per_it = []
            for rcd in item:
                ipltrcd = rcd.split('-')
                per_it.append(ipltrcd)
            sep_turn2nd.append(per_it)
        print(sep_turn2nd)

        #Chuyển số dạng chuỗi trở lại dạng số
        sep_turn3rd = []
        for item in sep_turn2nd:
            per_it = []
            for rcd in item:
                per_rcd = []
                for num in rcd:
                    per_rcd.append(float(num))
                per_it.append(per_rcd)
            sep_turn3rd.append(per_it)
        print(sep_turn3rd)


        #Tính tổng lương để trả theo profit
        dstlrated = []
        for person in range(0,len(sep_turn3rd)):
            total_ic_created = view_list[person][4]/1000000
            step_paid = [0]
            for counter in range(0,len(sep_turn3rd[person])):
                #print(sep_turn3rd[person][counter])
                if len(sep_turn3rd[person][counter]) == 3:
                    if total_ic_created > sep_turn3rd[person][counter][1]:
                        mul_amt = sep_turn3rd[person][counter][1]
                        total_ic_created -= sep_turn3rd[person][counter][1]
                    else:
                        mul_amt = total_ic_created
                        total_ic_created = 0

                    ip_paidstep = mul_amt * sep_turn3rd[person][counter][2]
                    step_paid.append(ip_paidstep)

             #Tính tổng profit cho trung tâm sau payment nhân viên   
            total_paid = sum(step_paid)
            print(str(total_ic_created) + ' - ' + str(step_paid) + ' -> ' + str(total_paid))
            dstlrated.append(total_paid*1000000)
        
        for item in range(0,len(dstlrated)):
            dstlrated[item] = int(dstlrated[item])

        print(dstlrated)
        
        #Xử lý profit
        ttpaid_alltea = []
        for person in range(0,len(dstlrated)):
            ttpaid_alltea.append(dstlrated[person]+view_list[person][3])
        print(ttpaid_alltea)

        pf_aftRated = []
        for person in range(0,len(ttpaid_alltea)):
            pf_aftRated.append(view_list[person][4]-ttpaid_alltea[person])

        #Push recs lên view
        for row in range(0,len(view_list)):
            for col in range(0,3):
                self.payment_view.setItem(row,col,QTableWidgetItem(str(view_list[row][col])))
            self.payment_view.setItem(row,3,QTableWidgetItem(str(self.viewnum(view_list[row][4]))))
            self.payment_view.setItem(row,5,QTableWidgetItem(str(self.viewnum(view_list[row][3]))))    
            self.payment_view.setItem(row,6,QTableWidgetItem(str(self.viewnum(ttpaid_alltea[row]))))    
            self.payment_view.setItem(row,4,QTableWidgetItem(str(self.viewnum(dstlrated[row]))))
            self.payment_view.setItem(row,7,QTableWidgetItem(str(self.viewnum(pf_aftRated[row]))))    



class OpWinM53_lctluonggv(Ui_chitraluonggv):
    
    #Def init
    def __init__(self):

        #Lọc danh sách mã giáo viên
        cmd_takedtb = 'SELECT maGV, tenGV FROM dsgiaovien;'
        threading.Thread(run_cmd(cmd_takedtb))
        self.mgv_list = []
        self.gv_list = []
        #print(recs)
        for item in recs:
            self.mgv_list.append(item[0])
            ipstr = item[0] + '-' + item[1]
            self.gv_list.append(ipstr)
        print(self.mgv_list)
        print(self.gv_list)

        #Lọc danh sách mã chi trả
        cmd_takelistpm = 'SELECT pmCode FROM traluonggv;'
        threading.Thread(run_cmd(cmd_takelistpm))
        print(recs)
        self.pmcodelist = []
        for item in recs:
            self.pmcodelist.append(str(item[0]))
        print(self.pmcodelist)

        #Lấy tháng, năm hiện tại
        now = datetime.datetime.now()
        print(now)
        self.month = now.month
        self.year = now.year
        print(self.month, '-', self.year)



    
    #Khởi tạo  
    def setupUi(self,window_53):
        super().setupUi(window_53)

        #Update dữ liệu mặc định tháng-năm
        self.thang.setValue(self.month)
        self.nam.setValue(self.year)
        self.thangtraluong.setValue(self.month)
        self.namtraluong.setValue(self.year)
        self.thangend.setValue(self.month)
        self.namend.setValue(self.year)


        #Define khi update giáo viên
        self.magv.currentIndexChanged.connect(self.update_cur_payment)


        #Cập nhật mã giáo viên lên list
        self.magv.addItems(self.gv_list)
        self.madel.addItems(self.pmcodelist)



        #Button clicked definition
        self.cndlhientai.clicked.connect(self.updateview)
        self.alldata.clicked.connect(self.updateallview)
        self.hoantat.clicked.connect(self.ql)
        self.traluong.clicked.connect(self.onpayment)
        self.xoamachitra.clicked.connect(self.delpayment)
        self.savedb.clicked.connect(self.savedata)
        self.createfile.clicked.connect(self.exportdata)

        #Def trạng thái đầu của button
        self.traluong.setEnabled(False)

        #Show current window
        window_53.show()

    #Def nút quay lại
    def ql(self):
        window_5.show()
        window_53.hide()

    #Def chuyển số integer thành chuỗi có dot seperate
    def viewnum(self,ipnum):
        strnum = str(ipnum)
        #print(strnum)
        liststr = []
        while len(strnum)>0:
            if len(strnum) >= 3:
                #print(strnum[-4:-1])
                liststr.append(strnum[len(strnum)-3:])
                strnum = strnum[:len(strnum)-3]
                #print(strnum)
                
            else:
                liststr.append(strnum)
                strnum = ''

        opstr = ''
        for item in liststr:
          opstr = item + '.' + opstr
        
        opstr = opstr[:-1]

        return opstr

    #Def update table view
    def updateview(self):

        self.mth = self.thang.value()
        self.yr = self.nam.value()
        #print(self.mth,'-',self.yr)
        #Cắt lọc đẩy lên table
        
        cmd_takefortb = 'SELECT traluonggv.pmCode, traluonggv.maGV, dsgiaovien.tenGV, traluonggv.PaidAmount, traluonggv.MonthPay, traluonggv.PaidDate FROM traluonggv INNER JOIN dsgiaovien ON dsgiaovien.maGV = traluonggv.maGV WHERE MONTH(traluonggv.MonthPay) = ' + str(self.mth) + ' AND YEAR(traluonggv.MonthPay) = ' + str(self.yr) + ';'
        #print(cmd_takefortb)
        threading.Thread(run_cmd(cmd_takefortb))
        print(recs)

        self.listupload = []
        for item in recs:
            peritem = []
            for it in item:
                peritem.append(it)
            self.listupload.append(peritem)

        print(self.listupload)

        #Tạo table
        self.chitra_view.setRowCount(len(self.listupload))
        self.chitra_view.setColumnCount(6)
        self.header20 = self.chitra_view.horizontalHeader()
        for col in range(0,6):
            self.header20.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Chèn dữ liệu lên table
        for row in range(0,len(self.listupload)):
            for col in range(0,6):
                if col ==  3:
                    self.chitra_view.setItem(row,col,QTableWidgetItem(str(self.viewnum(self.listupload[row][col]))))
                else:
                    self.chitra_view.setItem(row,col,QTableWidgetItem(str(self.listupload[row][col])))

        #Trả mả chi trả mới lên list xóa
        ipdelitems = []
        for item in self.listupload:
            ipdelitems.append(str(item[0]))

        #print(ipdelitems)
        self.madel.clear()
        self.madel.addItems(ipdelitems)

    #Def update all table view
    def updateallview(self):

        self.mth = self.thang.value()
        self.yr = self.nam.value()
        #print(self.mth,'-',self.yr)
        #Cắt lọc đẩy lên table
        
        cmd_takefortb = 'SELECT traluonggv.pmCode, traluonggv.maGV, dsgiaovien.tenGV, traluonggv.PaidAmount, traluonggv.MonthPay, traluonggv.PaidDate FROM traluonggv INNER JOIN dsgiaovien ON dsgiaovien.maGV = traluonggv.maGV;'
        #print(cmd_takefortb)
        threading.Thread(run_cmd(cmd_takefortb))
        print(recs)

        self.listupload = []
        for item in recs:
            peritem = []
            for it in item:
                peritem.append(it)
            self.listupload.append(peritem)

        print(self.listupload)

        #Tạo table
        self.chitra_view.setRowCount(len(self.listupload))
        self.chitra_view.setColumnCount(6)
        self.header20 = self.chitra_view.horizontalHeader()
        for col in range(0,6):
            self.header20.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Chèn dữ liệu lên table
        for row in range(0,len(self.listupload)):
            for col in range(0,6):
                if col ==  3:
                    self.chitra_view.setItem(row,col,QTableWidgetItem(str(self.viewnum(self.listupload[row][col]))))
                else:
                    self.chitra_view.setItem(row,col,QTableWidgetItem(str(self.listupload[row][col])))

        #Trả mả chi trả mới lên list xóa
        ipdelitems = []
        for item in self.listupload:
            ipdelitems.append(str(item[0]))

        #print(ipdelitems)
        self.madel.clear()
        self.madel.addItems(ipdelitems)

    #Def calc tổng lương đã trả
    def update_allpaid(self,ipmgv):
        thg = str(self.thangtraluong.value())
        nm = str(self.namtraluong.value())
        cmd_udluongdatra = 'SELECT SUM(PaidAmount) FROM traluonggv WHERE maGV = "' + ipmgv + '" AND MONTH(MonthPay)=' + thg + ' AND YEAR(MonthPay)=' + nm + ';'
        print(cmd_udluongdatra)
        run_cmd(cmd_udluongdatra)
        return recs[0][0]


    #Def calc tổng thu đã tạo ra cho trung tâm
    def update_totalcreated(self,ipmgv):
        thg = str(self.thangtraluong.value())
        nm = str(self.namtraluong.value())
        cmd_udtongthu = 'SELECT SUM(defhocphi.HocPhi) FROM thuhocphi INNER JOIN defhocphi ON defhocphi.IDhocphi = thuhocphi.IDhocphi WHERE defhocphi.MaGV = "' + ipmgv + '" AND MONTH(thuhocphi.NgayThuHP)=' + thg + ' AND YEAR(thuhocphi.NgayThuHP)=' + nm + ';'
        print(cmd_udtongthu)
        threading.Thread(run_cmd(cmd_udtongthu))
        return recs[0][0]

    #Def tổng lương phải trả theo rate
    def update_totalpaidformonth(self,ipmgv,iptotalcreated):
        cmd_rate = 'SELECT PayRate,BasePayment FROM dsgiaovien WHERE maGV="' + ipmgv + '";'
        threading.Thread(run_cmd(cmd_rate))
        calrate_str = recs[0][0]
        base_payment = recs[0][1]
        calrate_str1 = calrate_str.split(',')
        calrate_liststr = []
        for item in calrate_str1:
            peritem = item.split('-')
            calrate_liststr.append(peritem)

        calrate_list = []
        for item in calrate_liststr:
            peritem = []
            if len(item)>1:
                peritem.append(int(item[0]))
                peritem.append(float(item[1]))
                peritem.append(float(item[2]))
            else:
                peritem = [0,0,0]
            calrate_list.append(peritem)

        print(calrate_list,' - ',base_payment,' - ',type(base_payment))
        total_profit_created = float(iptotalcreated)
        luongrate = 0
        for item in calrate_list:
            if item[1]*1000000 <= total_profit_created:
                total_profit_created -= float(item[1]*1000000)
                luongrate += item[1]*item[2]*1000000
            elif total_profit_created > 0:
                luongrate += total_profit_created*item[2]
                total_profit_created = 0
        
        #print(self.viewnum(int(luongrate)))
        return int(luongrate)


    #Def tính lương và hiển thị lương cá nhân theo tháng hiện tại khi được chọn
    def update_cur_payment(self):
        thg = self.thangtraluong.value()
        nm = self.namtraluong.value()
        mgv_string= self.magv.currentText()
        mgv_split = mgv_string.split('-')
        mgv = mgv_split[0]
        print(mgv,'-',thg,'-',nm)

        #Lấy tổng lương
        totaloup = self.update_allpaid(mgv)
        if totaloup is not None:
            totalpaidinint = int(totaloup) 
        else:
            totalpaidinint = 0

        print(type(totalpaidinint),'-',totalpaidinint)
        #Update tổng lương đã trả lên view
        self.luongdatra.setText(self.viewnum(totalpaidinint))

        #Lấy tổng thu nhập tạo ra cho trung tâm
        total_incomecreated = self.update_totalcreated(mgv)
        if total_incomecreated is None:
            total_incomecreated = 0

        print(total_incomecreated)
        self.tongthu.setText(self.viewnum(total_incomecreated))

        #Lấy tổng thu nhập tính theo rate
        self.update_totalpaidformonth(mgv,total_incomecreated)
        oppayment = self.update_totalpaidformonth(mgv,total_incomecreated)
        print('Lương Rated:',oppayment)

        #Push lương Rate lên bảng chi
        self.luongrated.setText(self.viewnum(oppayment))
        
        #Lấy lương Base
        cmd_paybase = 'SELECT BasePayment FROM dsgiaovien WHERE maGV="' + mgv + '";'
        threading.Thread(run_cmd(cmd_paybase))
        op_paybase = recs[0][0]
        print('Lương cố định: ',op_paybase)
        #Push lương Base lên bảng chi
        self.luongbase.setText(self.viewnum(op_paybase))

        #Tính tổng lương và upload view
        totalpay = oppayment + op_paybase
        print('Tổng lương trong tháng: ',totalpay)
        self.tongluongthang.setText(self.viewnum(totalpay))

        #Tính tổng lương còn lại phải trả
        remainpaid = totalpay - totalpaidinint
        print('Tổng lương còn lại: ',remainpaid)
        self.luongphaitra.setText(self.viewnum(remainpaid))

        #Tính tổng lợi nhuận thuần tạo ra cho trung tâm
        totalprofit = total_incomecreated - totalpay
        print('Tổng lợi nhuận tạo ra: ',totalprofit)
        self.tonglai.setText(self.viewnum(totalprofit))

        #Set số tiền còn lại phải trả trên bảng view
        viewremain = round(remainpaid/1000000,2)
        print(viewremain)
        self.paidamount.setValue(viewremain)

        #Hiển thị lại nút trả lương để cho phép trả lương giáo viên hiện tại
        self.traluong.setEnabled(True)


    #Def nút tiến hành trả lương
    def onpayment(self):
        thg = self.thangtraluong.value()
        nm = self.namtraluong.value()
        mgv_string= self.magv.currentText()
        mgv_split = mgv_string.split('-')
        mgv = mgv_split[0]
        payment = self.paidamount.value()*1000000
        print(mgv,'-',payment,'-',thg,'-',nm)

        now = datetime.datetime.now()
        if nm >= 10:
            datenote = str(nm)+ '-' + str(thg) + '-01'
        else:
            datenote = str(nm)+ '-0' + str(thg) + '-01'

        print(now,' <-> ',datenote)

        #Cập nhật thông tin vào cơ sở dữ liệu
        cmd_takemaxid = 'SELECT MAX(pmCode) FROM traluonggv;'
        threading.Thread(run_cmd(cmd_takemaxid))
        maxid = recs[0][0] + 1
        #print(maxid)

        cmd_payment = 'INSERT INTO traluonggv (pmCode, maGV, PaidAmount, PaidDate, MonthPay) VALUES (' + str(maxid) + ', "' + str(mgv) + '", ' + str(payment) + ', "' + str(now) + '", "' + str(datenote) + '");'
        #print(cmd_payment)
        threading.Thread(run_cmd(cmd_payment))
        self.updateallview()

    #Def nút xóa mã chi trả
    def delpayment(self):
        ipdelid = self.madel.currentText()
        #print(ipdelid)
        cmd_del = 'DELETE FROM traluonggv WHERE pmCode = ' + str(ipdelid) + ';'
        print(cmd_del)
        threading.Thread(run_cmd(cmd_del))
        self.updateallview()

    #Def nút lưu dữ liệu lên cơ sở dữ liệu và vào bảng tongketluong
    def savedata(self):

        #Lưu dữ liệu lên cơ sở dữ liệu
        #Lấy toàn bộ mã giáo viên
        cmd_collectdata = 'SELECT maGV FROM dsgiaovien;'
        threading.Thread(run_cmd(cmd_collectdata))
        listmgv = []
        for item in recs:
            listmgv.append(item[0])
        #print(listmgv)

        #Ouput Mã giáo viên combomgv -> str_listmgv
        str_listmgv = str(listmgv)
        print(str_listmgv)


        #Lọc toàn bộ dữ liệu chi trả trong tháng
        #Lấy tháng và năm
        thg = self.thang.value()
        nm = self.nam.value()
        
        #Trả thông tin thangkettoan -> str_thang và namkettoan -> str_nam
        str_thang = str(thg)
        str_nam = str(nm)
        #print(str_thang + '/' + str_nam)


        #Id tháng năm để nhập dữ liệu
        if thg < 10:
            str_thg = '0' + str(thg)
        str_nam = str(nm)
        str_day = '01'

        #Tháng và năm tính lương để nhập vào cơ sở dữ liệu
        strdt = str_nam + '-' + str_thg + '-' + str_day
        #print(strdt)


        #Lọc toàn bộ dữ liệu trả lương trong tháng
        cmd_takedata = 'SELECT * FROM traluonggv WHERE MONTH(MonthPay)=' + str(thg) + ' AND YEAR(MonthPay)=' + str(nm) + ';'
        threading.Thread(run_cmd(cmd_takedata))
        #print(recs)
        allpm = []
        for item in recs:
            allpm.append(item)
        #print(allpm)

        #Lấy thông tin combopayment -> totalpm
        totalpm = []
        for id in listmgv:
            sum = 0
            for item in allpm:
                if item[1] == id:
                    sum += item[2]
            totalpm.append(sum)

        print(totalpm)
        

        #Tính toán tổng payment -> Hiển thị là totalpayment -> str_calc_all_pm , tính toán với str_calc_all_pm
        calc_all_pm = 0
        for item in totalpm:
            calc_all_pm += item
        #print(calc_all_pm)
        str_calc_all_pm = self.viewnum(calc_all_pm)
        print(str_calc_all_pm)



        #Lọc toàn bộ học phí đã thu trong tháng
        cmd_takeallmonthic = 'SELECT thuhocphi.IDthuHP, thuhocphi.IDhocphi, thuhocphi.NgayThuHP, defhocphi.HocPhi, defhocphi.MaGV, thuhocphi.MaHS  FROM thuhocphi INNER JOIN defhocphi ON defhocphi.IDhocphi = thuhocphi.IDhocphi WHERE MONTH(thuhocphi.NgayThuHP)= ' + str(thg) + ' AND YEAR(thuhocphi.NgayThuHP)= ' + str(nm) + ';'
        #print(cmd_takeallmonthic)
        threading.Thread(run_cmd(cmd_takeallmonthic))
        allic = []
        for item in recs:
            allic.append(item)
        #print(allic)

        #Xử lý tổng lãi ròng của mỗi giáo viên tạo thành chuỗi để lưu
        #List mã giáo viên theo string
        str_mgv = ','.join(listmgv)

        list_iccrt = []
        str_iccrt = ''
        totalic = 0
        for item in listmgv:
            sum = 0
            for it in allic:
                if it[4] == item:
                    #print(it)
                    sum += it[3]
            totalic += sum
            list_iccrt.append(sum)
        


        #List incomecreated: totalincome -> str_totalic và tính toán với totalic
        str_iccrt = str(list_iccrt)    
        #Lấy tổng incomes
        str_totalic = self.viewnum(totalic)

        #print(str_iccrt)
        #print(str_totalic)

        #Tính total_profit - totalprofit -> str_total_profit
        total_profit = totalic - calc_all_pm
        str_total_profit = self.viewnum(total_profit)

        #print(str_total_profit)

        #Nhập hoặc cập nhật dữ liệu lên database
        
        #Xử lý id key để nhập vào bảng lưu trữ kết toán hàng tháng
        cmd_importdata = 'SELECT idtongketluong FROM tongketluong WHERE thangkettoan = ' + str_thang + ' AND namkettoan = ' + str_nam + ';'
        #print(cmd_importdata)
        run_cmd(cmd_importdata)
        #print(recs)
        #print(len(recs))
        detect = len(recs)
        style = "Add"
        if detect !=0:
            oupdt = recs[0][0]
            style = "Alter"
        else:
            cmd_max = 'SELECT MAX(idtongketluong) FROM tongketluong;'
            run_cmd(cmd_max)
            oupdt = recs[0][0]

            if not(isinstance(recs[0][0],int)):
                oupdt = 1
            else:
                oupdt += 1

        #print(len(recs))
        #print(oupdt)

        #Chèn thông tin vào cơ sở dữ liệu:
        
        #Style == "Alter" hoặc "Add" (Quyết định chèn hoặc sửa thông tin)
        #idtongketluong == oupdt
        #combomagv == str_listmgv
        #combopayment == totalpm
        #thangkettoan == str_thang
        #namkettoan == str_nam
        #totalpayment == str_calc_all_pm
        #totalincome == str_totalic
        #totalprofit == str_total_profit

        if style == "Add":
            cmd_import = 'INSERT INTO tongketluong (idtongketluong,combomagv,combopayment,thangkettoan,namkettoan,totalpayment,totalincome,totalprofit) VALUES (' + str(oupdt) + ', "' + str_listmgv + '", "' + str(totalpm) + '", ' + str_thang + ', ' + str_nam + ', ' + str(calc_all_pm) + ', ' + str(totalic) + ', ' + str(total_profit) + ');'
            #print(cmd_import)
            run_cmd(cmd_import)
        elif style == "Alter":
            cmd_import = 'UPDATE tongketluong SET combomagv = "' + str_listmgv + '", combopayment = "' + str(totalpm) + '", thangkettoan = ' + str_thang + ', namkettoan = ' + str_nam + ', totalpayment = ' + str(calc_all_pm) + ', totalincome = ' + str(totalic) + ', totalprofit = ' + str(total_profit) + ' WHERE idtongketluong = ' + str(oupdt) + ';'
            print(cmd_import)
            run_cmd(cmd_import)







    #Tạo split tiền đề để lấy cụ thể mã giáo viên cho việc xuất
    def split_mgv(self,string_input):
        #Cắt rawdata mã giáo viên ra
        raw_data = string_input[1:len(string_input)-1]
        raw_data = raw_data.replace(' ','')
        #print(raw_data)
        cooking_1st_data = raw_data.split(',')
        

        #Cắt bỏ ngoặt trên từng mã giáo viên
        return_data = []
        for item in cooking_1st_data:
            per_item = item[1:len(item)-1]
            return_data.append(per_item)
        
        return return_data


    def split_payment_row(self,string_input):
        #Cắt rawdata payment row ra
        raw_data = string_input[1:len(string_input)-1]
        raw_data = raw_data.replace(' ','')
        cooked_data = raw_data.split(',')

        return cooked_data

    
    #Def chuyển đổi tháng năm để tiện so sánh trong điều kiện xuất dữ liệu
    def timetoint(self,ipmonth,ipyear):
        oup_int = ipyear * 100 + ipmonth
        return oup_int




    #Def nút đẩy dữ liệu ra file csv hoặc file excel từ bảng tổng kết lương
    def exportdata(self):
        thang_start = self.thangstart.value()
        nam_start = self.namstart.value()
        thg = self.thangend.value()
        nm = self.namend.value()
        print('Start: ', thang_start,'/',nam_start,'-',thg,'/',nm)

        #Lấy dữ liệu xuất workbook
        cmd_takedata = 'SELECT * FROM tongketluong WHERE (((namkettoan * 100) + thangkettoan) >= ' + str(self.timetoint(thang_start,nam_start)) + ') AND (((namkettoan * 100) + thangkettoan) <= ' + str(self.timetoint(thg,nm)) + ');'
        #print(cmd_takedata)
        run_cmd(cmd_takedata)
        #print(recs)
        
        export_data = []
        for item in recs:
            #print(item)
            per_row_data =[]
            for it in range(0,len(item)):
                per_row_data.append(item[it])
            
            #print(per_row_data)
            export_data.append(per_row_data)

        #print(export_data)
        
        #Tổng bộ thông tin xuất ra từ export_data
        #Cook rawdata to export


        cook_mgv = []
        set_mgv = set()
        for item in export_data:
            cook_mgv = item[1][1:len(item[1])-1]
            split_mgv = cook_mgv.split(',')
            #Check split_mgv
            #print(split_mgv)
            #for it in split_mgv:
            #    print(it)
            
            set_mgv = set_mgv | set(split_mgv)
        
        #print(set_mgv)

        #Mã giáo viên để xuất nằm ở list cook_mgv
        cook_mgv = list(set_mgv)
        print(cook_mgv)

        #Test lại mẫu xử lý kiểu dữ liệu set
        #set_test1 = set([1,2,3])
        #set_test2 = set([2,3,4])
        #set_oup = set()
        #set_oup = set_test1 | set_test2
        #print(set_oup)

        #Lấy dữ liệu giáo viên để xuất data -> Toàn bộ danh sách giáo viên
        cmd_infogv = 'SELECT maGV, tenGV FROM dsgiaovien;'
        run_cmd(cmd_infogv)

        dsgv = []
        for item in recs:
            peritem = list(item)
            dsgv.append(peritem)
      
        print(dsgv)

        #Xử lý thông tin lương trả từng giáo viên theo tháng -> all_payment_rows
        all_payment_rows = []


        #Mỗi dòng tương ứng với một tháng, các cột là toàn bộ giáo viên trong danh sách giáo viên đã tạo (dsgv)
        for item in export_data:
            per_row = []

            #Append idtongketluong, thangkettoan, namkettoan vào trước
            per_row.append(item[0])
            per_row.append(item[3])
            per_row.append(item[4])

            #Tạo tháng năm kết hợp
            if len(str(item[3]))>1:
                thangnam = str(item[4]) + str(item[3])
            else:
                thangnam = str(item[4]) + '0' + str(item[3])
            per_row.append(thangnam)

            #Lấy string combomgv và combopayment của item này
            str_cbmgv = item[1]
            str_cbpm = item[2]

            list_cbmgv = self.split_mgv(str_cbmgv)
            list_cbpm = self.split_payment_row(str_cbpm)
            #print(list_cbmgv)
            #print(list_cbpm)


            #Vấn đề tiếp theo ở đây là lấy từng row theo đúng mã giáo viên ở từng cột đã lưu ở dsgv
            for gv in dsgv:
                #Lấy index giáo viên được chọn hiện tại trong dsgv
                gv_index = list_cbmgv.index(gv[0])
                
                #Dưa payment vào per_row xuất
                per_row.append(list_cbpm[gv_index])
                #print(list_cbpm[gv_index])

            #Chèn tiếp tất toán và các thông tin liên quan của tháng này vào combo payment
            per_row.append(item[5])
            per_row.append(item[6])
            per_row.append(item[7])                

            all_payment_rows.append(per_row)

        print(all_payment_rows)

        #Danh sách giáo viên (Xuất 2 dòng đầu bao gồm mã giáo viên và tên giáo viên): dsgv
        #Các data gốc bao gồm: idtongketluong, tháng, năm, namthang, payment ứng với từng giáo viên, totalpayment, totalincome, totalprofit -> all_payment_rows


        #Tạo workbook để export dữ liệu
        wb = Workbook()
        #Tạo sheet mới và set ở vị trí đầu tiên
        dest_filename = 'KhanhVuEdu_PaymentData.xlsx'
        ws = wb.active
        
        ws_title = str(thang_start) + '_' + str(nam_start) + '<to>' + str(thg) + '_' + str(nm)
        print(ws_title)
        ws.title = ws_title

        #Viết dữ liệu vào file
        #Lệnh viết: sheet.cell(row=2,column=2).value = 2
        #Lệnh lưu: book.save('write2cell.xlsx')

        #Khởi tạo để xuất line đầu tiên
        sheet_row = 1
        sheet_col = 4

        for teacher in dsgv:
            sheet_col += 1
            ws.cell(row=sheet_row,column=sheet_col).value = teacher[1]
            sheet_row += 1
            ws.cell(row=sheet_row,column=sheet_col).value = teacher[0]
            sheet_row -= 1

        #Trả sheet_col, sheet_row để xử lý xuất payment (trả ngược một đơn vị để tiện inc)
        sheet_row = 2
        sheet_col = 0

        for item in all_payment_rows:
            print(item)

            sheet_row += 1
            for it in item:
                sheet_col += 1

                #Xuất dữ liệu vào file excel
                ws.cell(row=sheet_row,column=sheet_col).value = int(it)

            col_numsave = sheet_col
            sheet_col = 0

        #Trả dữ liệu rời vào file gốc này
        ws.cell(row=2,column=1).value = 'ID Kết toán'
        ws.cell(row=2,column=2).value = 'Tháng'
        ws.cell(row=2,column=3).value = 'Năm'
        ws.cell(row=1,column=4).value = 'Tên giáo viên'
        ws.cell(row=2,column=4).value = 'Mã giáo viên'

        #Tạo tiền đề lưu tiêu đề kết tháng
        sheet_row = 1
        sheet_col = col_numsave
        print(sheet_row, ' - ',sheet_col)

        #Lưu 3 cột đầu vào
        ws.cell(row=sheet_row,column=sheet_col-2).value = "Tổng lương"
        ws.cell(row=sheet_row,column=sheet_col-1).value = "Tổng lãi ròng"
        ws.cell(row=sheet_row,column=sheet_col).value = "Tổng lợi tức sau lương"

        #ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4) (totalpayment, totalincome, totalprofit) -> 3 cột cuối
        ws.merge_cells(start_row=1,start_column=sheet_col, end_row=2,end_column=sheet_col)
        ws.merge_cells(start_row=1,start_column=sheet_col-1, end_row=2,end_column=sheet_col-1)
        ws.merge_cells(start_row=1,start_column=sheet_col-2, end_row=2,end_column=sheet_col-2)





        #Lưu file
        wb.save(dest_filename)





        #Grab the active worksheet
        #ws = wb.active
        #Data can be assigned directly to cells
        #ws['A1'] = 42




#This part is intended left blank


#Window 6 - Quản lý bài học và giờ học
class OpWinM6_qlbhgh(Ui_quanlybaihocgiohoc):

    #Khởi tạo
    def setupUi(self,window_6):
        super().setupUi(window_6)

        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)

        #Show current window
        window_6.show()

    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_6.hide()









# These spaces are intended left blank












#Window 7 - Quản lý thông báo phụ huynh online
class OpWinM7_ttphonline(Ui_M7_thongtinlienhephuhuynh):
   
    #Khởi tạo
    def setupUi(self,window_7):
        super().setupUi(window_7)

        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.themxoacapnhat.clicked.connect(self.txcn)
        self.lhhsdonle.clicked.connect(self.gtncn)
        self.lhhstheonhom.clicked.connect(self.gtntn)

        #Show current window
        window_7.show()

    #Def nút quay lại
    def ql(self):
        MainWinOpt.show()
        window_7.hide()

    #Def nút thêm xóa cập nhật thông tin phụ huynh
    def txcn(self):
        self.ui = OpWinM71_txcninfoph()
        self.ui.setupUi(window_71)
        window_7.hide()

    #Def nút mở window gửi tin nhắn tới học sinh riêng lẻ
    def gtncn(self):
        self.ui = OpWinM72_guitncn()
        self.ui.setupUi(window_72)
        window_7.hide()

    #Def nút mở window gửi tin nhắn học sinh theo nhóm
    def gtntn(self):
        self.ui = OpWinM73_lhhstn()
        self.ui.setupUi(window_73)
        window_7.hide()


class OpWinM71_txcninfoph(Ui_themxoacapnhatinfophuhuynh):
    #Chạy nền trước khởi tạo
    def __init__(self):
        #Lọc mã nhóm
        cmd_laynhom = 'SELECT MaNhom FROM dsnhom;'
        run_cmd(cmd_laynhom)
        #print(recs)
        self.listnhom = ['0']
        for row in recs:
            self.listnhom.append(row[0])
        #print(self.listnhom)

        #Lọc mã học sinh & Tên học sinh từ bảng onlineconnection
        cmd_lochs = 'SELECT onlineconnection.MaHS, dshocsinh.HoTenHS FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS;'
        self.stucodes = []
        self.stunames = []
        run_cmd(cmd_lochs)
        for row in recs:
            self.stucodes.append(row[0])
            self.stunames.append(row[1])

        #print(self.stucodes)
        #print(self.stunames)
        self.stuinfos = []
        for row in range(0,len(self.stucodes)):
            rcdip = self.stucodes[row] + ' - ' + self.stunames[row]
            self.stuinfos.append(rcdip)

        print(self.stuinfos)

        #Lọc toàn bộ mã học sinh và tên học sinh từ danh mục học sinh
        cmd_locallhs = 'SELECT MaHS, HoTenHS FROM dshocsinh;'
        self.allstucodes = []
        self.allstunames = []
        run_cmd(cmd_locallhs)
        for row in recs:
            self.allstucodes.append(row[0])
            self.allstunames.append(row[1])



    #Khởi tạo
    def setupUi(self, window_71):
        super().setupUi(window_71)

        #Init create listnhom
        self.nhom.addItems(self.listnhom)
        self.mahsdelinfoph.addItems(self.stuinfos)

        #Button clicked definition
        self.quaylai.clicked.connect(self.ql)
        self.loc.clicked.connect(self.but_loc)
        self.lapbangthem.clicked.connect(self.createtb)
        self.but_theminfo.clicked.connect(self.theminfo)
        self.capnhat.clicked.connect(self.updateinfo)
        self.delinfo.clicked.connect(self.xoattll)

        #Show current window
        window_71.show()


    #Def nút quay lại
    def ql(self):
        window_7.show()
        window_71.hide()

    #Def but lọc
    def but_loc(self):
        nhomcs = self.nhom.currentText()
        lopcs = self.lop.value()
        print(nhomcs + ' - ' + str(lopcs))

        #Lấy thông tin từ database
        if nhomcs != '0' and lopcs != 0:
            cmd_loc = 'SELECT onlineconnection.MaHS, onlineconnection.FBacclist, onlineconnection.FBIBMessengerList,dshocsinh.HoTenHS FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS INNER JOIN hstheonhom ON hstheonhom.MaHS = onlineconnection.MaHS WHERE dshocsinh.Lop = ' + str(lopcs) + ' AND hstheonhom.Manhom = "' + nhomcs + '";'
        elif nhomcs != '0' and lopcs == 0:
            cmd_loc = 'SELECT onlineconnection.MaHS, onlineconnection.FBacclist, onlineconnection.FBIBMessengerList,dshocsinh.HoTenHS FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS INNER JOIN hstheonhom ON hstheonhom.MaHS = onlineconnection.MaHS WHERE hstheonhom.Manhom = "' + nhomcs + '";'
        elif nhomcs == '0' and lopcs != 0:
            cmd_loc = 'SELECT onlineconnection.MaHS, onlineconnection.FBacclist, onlineconnection.FBIBMessengerList,dshocsinh.HoTenHS FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS INNER JOIN hstheonhom ON hstheonhom.MaHS = onlineconnection.MaHS WHERE dshocsinh.Lop = ' + str(lopcs) + ';'
        else:
            cmd_loc = 'SELECT onlineconnection.MaHS, onlineconnection.FBacclist, onlineconnection.FBIBMessengerList,dshocsinh.HoTenHS FROM onlineconnection INNER JOIN dshocsinh ON dshocsinh.MaHS = onlineconnection.MaHS;'
        
        run_cmd(cmd_loc)
        print(cmd_loc)

        opdata = []
        for row in recs:
            opdata.append(row)

        print(opdata)

        #Chuẩn bị bảng xuất dữ liệu lên view
        while self.viewinfo.rowCount() >0:
            self.viewinfo.removeRow(0)

        self.viewinfo.setRowCount(len(opdata))
        self.viewinfo.setColumnCount(4)
        self.header10 = self.viewinfo.horizontalHeader()
        for col in range(0,4):
            self.header10.setSectionResizeMode(col,QHeaderView.ResizeToContents)


        for row in range(0,len(opdata)):
            self.viewinfo.setItem(row,0,QTableWidgetItem(opdata[row][0]))
            self.viewinfo.setItem(row,1,QTableWidgetItem(opdata[row][3]))


            if opdata[row][1] is None:
                self.viewinfo.setItem(row,2,QTableWidgetItem('No Infomation!!!'))
            else:
                self.viewinfo.setItem(row,2,QTableWidgetItem(opdata[row][1]))

            
            if opdata[row][2] is None:
                self.viewinfo.setItem(row,3,QTableWidgetItem('No Infomation!!!'))
            else:
                self.viewinfo.setItem(row,3,QTableWidgetItem(opdata[row][2]))



    #Def nút lập bảng thêm thông tin
    def createtb(self):
        #Tạo bảng thêm thông tin
        while self.viewinfo.rowCount() >0:
            self.viewinfo.removeRow(0)

        self.amountrec = self.numinfoadd.value()
        print(self.amountrec)
        self.viewinfo.setRowCount(self.amountrec)
        self.viewinfo.setColumnCount(4)
        self.header9 = self.viewinfo.horizontalHeader()

        for col in range(0,4):
            self.header9.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        for row in range(0,self.amountrec):
            #Chèn Combobox vào bảng
            print(self.allstucodes)
            mahs_cb = QComboBox()
            mahs_cb.addItems(self.allstucodes)
            self.viewinfo.setCellWidget(row,0,mahs_cb)
       


    #Def nút cập nhật thông tin lên cơ sở dữ liệu
    def updateinfo(self):

        #Lấy ngược dữ liệu vào
        self.ipdata = []
        for row in range(0,self.amountrec):
            item = self.viewinfo.cellWidget(row,0)
            if isinstance(item,QComboBox):
                cur_mahs = item.currentText()
            per_rec = []
            per_rec.append(cur_mahs)
            self.ipdata.append(per_rec)

        print(self.ipdata)


        #Truy xuất thông tin tên học sinh
        for row in range(0,len(self.ipdata)):
            cmd_udstunames = 'SELECT HoTenHS FROM dshocsinh WHERE MaHS = "' + self.ipdata[row][0] + '";'
            run_cmd(cmd_udstunames)
            self.ipdata[row].append(recs[0][0])
        print(self.ipdata)

        #Load ngược lên bảng data
        for row in range(0,self.amountrec):
            self.viewinfo.setItem(row,1,QTableWidgetItem(self.ipdata[row][1]))


    #Def nút Thêm thông tin vào cơ sở dữ liệu
    def theminfo(self):
        #print(self.ipdata)
        #Tạo record thông tin
        for row in range(0,self.amountrec):
            #Cập nhật các col 2,3 vào self.ipdata
            item1 = self.viewinfo.item(row,2)
            item2 = self.viewinfo.item(row,3)
            if item1 and item1.text():
                self.ipdata[row].append(str(item1.text()))
            else:
                self.ipdata[row].append('')
            if item2 and item2.text():
                self.ipdata[row].append(str(item2.text()))
            else:
                self.ipdata[row].append('')

        print(self.ipdata)

        #Push lệnh đẩy vào cơ sở dữ liệu
        #Kiểm tra tính đẩy đủ dữ liệu của bảng onlineconnection
        cmd_onlrecs = 'SELECT MaHS FROM onlineconnection;'
        run_cmd(cmd_onlrecs)
        mhs_onl = []
        for row in recs:
            mhs_onl.append(row[0])

        cmd_dshsrecs = 'SELECT MaHS FROM dshocsinh;'
        run_cmd(cmd_dshsrecs)
        mhs_dshs = []
        for row in recs:
            mhs_dshs.append(row[0])

        diff = list(set(mhs_dshs) - set(mhs_onl))
        print(diff)
        print(len(diff))

        if len(diff) > 0:
            for row in diff:
                cmd_ipdata = 'INSERT INTO onlineconnection (MaHS,dshocsinh_MaHS) VALUES ("' + row + '", "' + row + '");'
                run_cmd(cmd_ipdata)
                print(cmd_ipdata)
        
        #Tiến hành rút dữ liệu và cập nhật lên database
        for row in self.ipdata:
            #Lấy dữ liệu trên database
            cmd_takedata = 'SELECT FBacclist, FBIBMessengerList FROM onlineconnection WHERE MaHS = "' + row[0] + '";'
            run_cmd(cmd_takedata)
            print(recs)

            stralter1 = ''
            stralter2 = ''
            
            if row[2] != '':
                if (recs[0][0] is None) or (recs[0][0] == ''):
                    stralter1 = row[2]
                else:
                    stralter1 = recs[0][0] + ',' + row[2]

            if row[3] != '':
                if (recs[0][1] is None)  or (recs[0][1] == ''):
                    stralter2 = row[3]
                else:
                    stralter2 = recs[0][1] + ',' + row[3]
            #Đang bị lỗi đoạn trên dưới đây do biến stralter không thỏa yêu cầu khi nhập chỉ FBacclist mà không nhập FBIBMessengerList
            if stralter1 == '':
                stralter1 = recs[0][0]
            if stralter2 == '':
                stralter2 = recs[0][1]

            print(type(stralter1))
            print(type(stralter2))
            print(stralter1)
            print(stralter2)
            if stralter1 is None:
                stralter1 = ''
            if stralter2 is None:
                stralter2 = ''

            cmd_updateinfo = 'UPDATE onlineconnection SET FBacclist = "' + stralter1 + '", FBIBMessengerList = "' + stralter2 + '" WHERE MaHS = "' + row[0] + '";'
            print(cmd_updateinfo)
            run_cmd(cmd_updateinfo)

    #Def nút xóa thông tin liên lạc trên cơ sở dữ liệu
    def xoattll(self):
        mahsdelinfo = self.mahsdelinfoph.currentText()
        
        while mahsdelinfo[len(mahsdelinfo)-1] != '-':
            #print(mahsdelinfo[len(mahsdelinfo)-1]) 
            mahsdelinfo = mahsdelinfo[:(len(mahsdelinfo)-1)]
            #print(mahsdelinfo[len(mahsdelinfo)-1]) 
        
        mahsdelinfo = mahsdelinfo[:(len(mahsdelinfo)-2)]
        print(mahsdelinfo)

        cmdxoa = 'UPDATE onlineconnection SET FBacclist = "", FBIBMessengerList = "" WHERE MaHS = "' + mahsdelinfo + '";'
        print(cmdxoa)
        run_cmd(cmdxoa)



#Def chức năng gửi tin nhắn đến account ID
def inboxfbacc(msg,idsend):
    #Chương trình gửi qua account ID
    
    client = Client("0353095909", "thptquocgia")
    client.send(Message(text=msg), thread_id=idsend, thread_type=ThreadType.USER)
    client.logout()
#Def chức năng gửi tin nhắn tới Messenger Inbox ID
def inboxfbmsg(msg,idsend):
    #Chương trình gửi qua account ID
    
    client = Client("0353095909", "thptquocgia")
    client.send(Message(text=msg), thread_id=idsend, thread_type=ThreadType.GROUP)
    client.logout()


#Def chức năng kiểm tra ID
def checkid(idip):
    lstip = ['0','1','2','3','4','5','6','7','8','9']
    op = True
    #print(idip)
    #print(type(idip))
    #print(idip[0])
    for i in range(0,len(idip)):
        if idip[i] not in set(lstip):
            op = False
    return op

#Def chức năng gửi tin nhắn đến cá nhân
class OpWinM72_guitncn(Ui_guitncn):
    #Init trước khi chạy window 72
    def __init__(self):
        cmd_mahs = 'SELECT MaHS FROM onlineconnection;'
        run_cmd(cmd_mahs)
        self.mhsop = []
        for row in recs:
            self.mhsop.append(row[0])


    #Khởi tạo
    def setupUi(self, window_72):
        super().setupUi(window_72)

        #Khởi tạo combobox
        self.mahschoice.addItems(self.mhsop)

        #Chạy khi update option
        self.mahschoice.currentIndexChanged.connect(self.updif)

        #Định nghĩa buttons
        self.quaylai.clicked.connect(self.ql)
        self.sendmess.clicked.connect(self.sdm)        

        #Hiển thị Window 72
        window_72.show()
    
    #def but quay lại
    def ql(self):
        window_7.show()
        window_72.hide()
    
    #Def khi cập nhật lựa chọn
    def updif(self):
        mhs = self.mahschoice.currentText()
        cmdtakeinfo = 'SELECT dshocsinh.HoTenHS, onlineconnection.fBacclist, onlineconnection.FBIBMessengerList FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS WHERE onlineconnection.MaHS = "' + mhs + '";'
        run_cmd(cmdtakeinfo)
        if (recs[0][1] is not None) and (recs[0][1] != ''):
            self.accid = recs[0][1]
        else: self.accid = ''
        if (recs[0][2] is not None) and (recs[0][2] != ''):
            self.msgid = recs[0][2]
        else: self.msgid = ''
        self.stuname.setText(recs[0][0])
        self.fbaccid.setText(recs[0][1])
        self.fbmessid.setText(recs[0][2])


    #Def khi gửi tin nhắn
    def sdm(self):
        startstr = ' - KhanhVuEducation Center - \n - Tin nhắn gửi đến bạn ' + self.stuname.text() + ': \n'
        endstr = 'Lưu ý: Vui lòng trả lời tin nhắn để xác nhận bạn đã nhận được tin. Reply nếu có bất kì yêu cầu và thắc mắc.'
        messop = self.mess.toPlainText()
        fullmess = startstr + '\n' + messop + '\n \n' + endstr
        print(fullmess)

        #print(checkid(str(self.accid)))
        #print(checkid(str(self.fbmessid)))
            
        if checkid(str(self.accid)):
            inboxfbacc(fullmess,self.accid)
        if checkid(str(self.fbmessid)):
            inboxfbacc(fullmess,self.fbmessid)

#Def chức năng liên hệ học sinh theo nhóm
class OpWinM73_lhhstn(Ui_gtnnhom):
    #Init section
    def __init__(self):
        cmdipnhom = 'SELECT MaNhom FROM dsnhom;'
        run_cmd(cmdipnhom)

        self.ipnhom = []
        for row in recs:
            self.ipnhom.append(row[0])

        


    #Khởi tạo
    def setupUi(self, window_73):
        super().setupUi(window_73)

        #Dồn nhóm vào combobox
        self.nhom.addItems(self.ipnhom)

        #Khởi tạo disable khung chọn học sinh để gửi
        self.mahsib.setDisabled(True)

        #Set disable nút gửi tin nhắn
        self.send.setEnabled(False)
        
        #Def onupdate nhóm
        self.nhom.currentTextChanged.connect(self.nhomcs)
        self.sdsome.toggled.connect(self.enableoption)
        self.sdall.toggled.connect(self.disableoption)

        #Def nút nhấn
        self.quaylai.clicked.connect(self.ql)
        self.send.clicked.connect(self.sendmessage)


        #Hiện window
        window_73.show()

    #def nút quay lại
    def ql(self):
        window_7.show()
        window_73.hide()

    #Def option chọn nhóm
    def nhomcs(self):
        cmdip = 'SELECT onlineconnection.MaHS, dshocsinh.HoTenHS, hstheonhom.MaNhom, onlineconnection.FBacclist, onlineconnection.FBIBMessengerList FROM onlineconnection INNER JOIN dshocsinh ON onlineconnection.MaHS = dshocsinh.MaHS INNER JOIN hstheonhom ON hstheonhom.MaHS = onlineconnection.MaHS WHERE hstheonhom.MaNhom = "' + self.nhom.currentText() + '";'
        print(cmdip)
        run_cmd(cmdip)
        print(recs)
        self.opstus = []
        for row in recs:
            perrow = []
            for item in row:
                perrow.append(item)
            self.opstus.append(perrow)

        #Chuẩn bị bảng để push thông tin lên
        self.hstrongnhom.setRowCount(len(self.opstus))
        self.hstrongnhom.setColumnCount(5)
        self.header12 = self.hstrongnhom.horizontalHeader()
        for col in range(0,5):
            self.header12.setSectionResizeMode(col,QHeaderView.ResizeToContents)

        #Push items lên table
        for row in range(0,len(self.opstus)):
            self.hstrongnhom.setItem(row,0,QTableWidgetItem(self.opstus[row][0]))
            self.hstrongnhom.setItem(row,1,QTableWidgetItem(self.opstus[row][1]))
            self.hstrongnhom.setItem(row,2,QTableWidgetItem(str(self.opstus[row][2])))
            self.hstrongnhom.setItem(row,3,QTableWidgetItem(str(self.opstus[row][3])))
            self.hstrongnhom.setItem(row,4,QTableWidgetItem(str(self.opstus[row][4])))

    #Def khi thay đổi radiobutton chọn gửi cho nhóm hay cho một số học sinh trong nhóm
    def enableoption(self):
        self.mahsib.setDisabled(False)
        self.send.setEnabled(True)
    
    def disableoption(self):
        self.mahsib.setDisabled(True)
        self.send.setEnabled(True)

    #Def đổi weekday thành thứ để gửi tin nhắn
    def chg_wd(self,ipint):
        if ipint == 0:
            return 'Thứ hai'
        if ipint == 1:
            return 'Thứ ba'
        if ipint == 2:
            return 'Thứ tư'
        if ipint == 3:
            return 'Thứ năm'
        if ipint == 4:
            return 'Thứ sáu'
        if ipint == 5:
            return 'Thứ bảy'
        if ipint == 6:
            return 'Chủ nhật'
    

    #Def nút gửi tin nhắn đến nhóm
    def sendmessage(self):
        #Xử lý các options checkbox
        lastdateck = self.baodiemdanh.isChecked()
        newsche = self.baolich.isChecked()
        feeck = self.baohocphi.isChecked()
        
        if self.sdall.isChecked():
            sta_all = True
            cmd_ip = 'SELECT MaHS FROM hstheonhom WHERE MaNhom = "' + self.nhom.currentText() + '";'
            run_cmd(cmd_ip)
            inputcodes = []
            for row in recs:
                inputcodes.append(row[0])

            print(sta_all)
            print(inputcodes)
        else:
            sta_all = False
            inputcodes = []
            inputcodes = self.mahsib.text().split(',')
            print(sta_all)
            print(inputcodes)

        #Lấy tên học sinh
        if inputcodes != []:
            inputnames = []
            for idhs in inputcodes:
                cmd_getname = 'SELECT HoTenHS FROM dshocsinh WHERE MaHS = "' + idhs + '";'
                run_cmd(cmd_getname)
                inputnames.append(recs[0][0])
            print(inputnames)

        for ctrec in range(0,len(inputcodes)):
            #Cấu trúc tin nhắn
            msg = ''
            startstr = ' - KhanhVuEducation Center - \n \n - Tin nhắn gửi đến bạn ' + inputnames[ctrec] + ': \n\n' 
            endstr = 'Lưu ý: Vui lòng trả lời tin nhắn để xác nhận bạn đã nhận được tin. Reply nếu có bất kì yêu cầu và thắc mắc.'
            
            msg += 'Tin nhắn gửi đến học sinh trong nhóm ' + str(self.nhom.currentText()) + ' \n \n'

            if lastdateck:
                print("Absent check activated!!")
                cmd_checklastlearn = 'SELECT NgayGioHoc FROM diemdanhhs WHERE MaHS = "' + inputcodes[ctrec] + '" AND MaNhom = "' + self.nhom.currentText() + '" ORDER BY IDDiemDanh DESC LIMIT 1;'
                print(cmd_checklastlearn)
                run_cmd(cmd_checklastlearn)
                opdt = 'None'
                if len(recs)>0:
                    opdt = recs[0][0]
                    print(opdt)
                #Cấu trúc tin nhắn
                msg_diemdanh = 'Buổi học cuối của bạn vào ngày: ' + str(opdt) + '\n \n'


                cmd_checkbuoihocmoinhat = 'SELECT NgayGioHoc FROM diemdanhhs WHERE MaNhom = "' + self.nhom.currentText() + '" ORDER BY IDDiemDanh DESC LIMIT 1;'
                run_cmd(cmd_checkbuoihocmoinhat)
                newestck = 'None'
                if len(recs)>0:
                    newestck = recs[0][0]
                    print(newestck)
                #Cấu trúc tin nhắn
                msg_diemdanh += 'Buổi học mới nhất của nhóm đã diễn ra vào ngày: ' + str(newestck) + ' \n \n'
                #print(msg_diemdanh)

                #Cập nhật vào tin nhắn chung
                msg += msg_diemdanh


            if newsche:
                print('New schedule activated!!')
                cmd_laytkb = 'SELECT TIME(NgayGioHoc), Thu FROM thoikhoabieu WHERE MaNhom = "' + self.nhom.currentText() + '" AND TinhTrangLich = 1;' 
                print(cmd_laytkb)
                run_cmd(cmd_laytkb)
                #print(recs)
                op_tkb = []
                if len(recs)>0:
                    for row in recs:
                        perrow = []
                        perrow.append(str(row[0]))
                        perrow.append(self.chg_wd(row[1]))
                        op_tkb.append(perrow)
                    print(op_tkb)
                    
                    #Tạo tin nhắn thời khóa biểu
                    msg_tkb = 'Thời khóa biểu hiện hành: '
                    for row in op_tkb:
                        msg_tkb += row[1] + ' ' + row[0] + ';   '
                    
                    #print(msg_tkb)
                    msg_tkb += '\n \n'

                    msg += msg_tkb
                
                    


            if feeck:
                msg_fee = ''
                print('Fee check activated!!')
                cmd_fee = 'SELECT SoBuoiMoiLanDongTien, SoLanDaDongTien, SoBuoiDaHoc, HocPhi FROM defhocphi INNER JOIN dsnhom ON dsnhom.MaMonHoc = defhocphi.MaMonHoc AND dsnhom.maGV = defhocphi.MaGV WHERE dsnhom.MaNhom = "' + self.nhom.currentText() +'" AND defhocphi.MaHS = "' + inputcodes[ctrec] + '";'
                print(cmd_fee)
                run_cmd(cmd_fee)
                print(recs)
                paid_cir = recs[0][0]
                paid_times = recs[0][1]
                studied = recs[0][2]
                cir_amount = recs[0][3]
                notpaid = studied - (paid_cir * paid_times)
                if notpaid < 0:
                    notpaid = 0

                #Xử lý message học phí
                msg_fee += 'Số buổi đã học ở trung tâm: ' + str(studied) + '\n'
                msg_fee += 'Số buổi/tháng quy định: ' + str(paid_cir) + '\n'
                msg_fee += 'Số lượt học phí đã đóng: ' + str(paid_times) + '\n'
                
                if notpaid >= paid_cir:
                    pmt = notpaid/paid_cir
                    msg_fee += 'Vui lòng đóng ' + str(int(pmt)) + ' tháng học phí. Tổng: ' + str(int(pmt*cir_amount)) + '\n \n'
                    cmd_laylistbuoi = 'SELECT NgayGioHoc FROM diemdanhhs WHERE MaHS = "' + inputcodes[ctrec] + '" AND MaNhom = "' + self.nhom.currentText() + '" ORDER BY IDDiemDanh DESC LIMIT '+ str(notpaid) + ';'
                    run_cmd(cmd_laylistbuoi)
                    listbuoi = []
                    for row in recs:
                        listbuoi.append(row[0])
                    oplistbuoi = ', '.join(map(str,listbuoi))
                    msg_fee += 'Danh sách buổi học: ' + oplistbuoi + '\n \n'




                if (notpaid == 7) or (notpaid == 6):
                    msg_fee += 'Lưu ý bạn sắp đến kỳ đóng học phí!! \n \n'

                #print(msg_fee)
                msg += msg_fee

                #Tin nhắn chung cho nhóm
                msg_common = 'TIN NHẮN CHUNG: \n \n' + self.message.toPlainText() 

                #In kiểm tra tin nhắn
                msg += msg_common + '\n \n'
                print(msg)



            #Hoàn tất đầu cuối tin nhắn
            msg = startstr + msg + endstr

            #Tiến hành gửi tin nhắn
            cmd_layid = 'SELECT FBacclist, FBIBMessengerList FROM onlineconnection WHERE MaHS = "' + inputcodes[ctrec] + '";'
            #print(cmd_layid)
            run_cmd(cmd_layid)
            print(recs)
            acclist = []
            messlist = []
            for row in recs:
                if row[0] != '':
                    listitems0 = []
                    listitems0 = row[0].split(',')
                    for itm0 in listitems0:
                        acclist.append(int(itm0))
                if row[1] != '':
                    listitems1 = []
                    listitems1 = row[1].split(',')
                    for itm1 in listitems1:
                        messlist.append(int(itm1))



            if len(acclist) >0:
                for id in acclist:
                    inboxfbacc(msg,id)

            if len(messlist) >0:
                for id in messlist:
                    inboxfbmsg(msg,id)

















#This part is intended left blank




#Validation controllers

#Must create Validation controllers here

#Main function program
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    
    #Variables Declaire for Windows Transitions
    window_1 = QDialog()
    window_11 = QDialog()
    window_12 = QDialog()
    window_13 = QDialog()
    window_14 = QDialog()
    window_16 = QDialog()



    window_2 = QDialog()
    window_21 = QDialog()
    window_22 = QDialog()
    window_23 = QDialog()
    window_24 = QDialog()


    window_3 = QDialog()
    window_31 = QDialog()
    window_32 = QDialog()
    window_33 = QDialog()

    window_4 = QDialog()
    window_41 = QDialog()
    window_42 = QDialog()
    window_43 = QDialog()

    window_5 = QDialog()
    window_51 = QDialog()
    window_52 = QDialog()
    window_53 = QDialog()

    window_6 = QDialog()
    

    window_7 = QDialog()
    window_71 = QDialog()
    window_72 = QDialog()
    window_73 = QDialog()

    #windows_loading = QDialog()
    



    MainWinOpt = QMainWindow()
    ui = MainWinOptimized()
    ui.setupUi(MainWinOpt)
    MainWinOpt.show()
    sys.exit(app.exec_())
    